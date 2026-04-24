import json
import pytest
from pathlib import Path
import tempfile
from tn5verify.refs import load_folder
from tn5verify.grouping import build_groups
from tn5verify.characterize import annotate
from tn5verify.target import build_multi_fasta
from tn5verify.align import run_bwa
from tn5verify.identify import classify
from tn5verify.integrity import evaluate
from tn5verify.render.json_dump import write as write_json, load as load_json
from tn5verify.render.well_report import write as write_report
from tn5verify.render.plate import write_xlsx

FIXTURES = Path(__file__).parent / "fixtures"
REFS = FIXTURES / "refs"
FASTQS = FIXTURES / "fastqs"

@pytest.fixture(scope="module")
def full_results(tmp_path_factory):
    tmp = tmp_path_factory.mktemp("m5")
    refs = load_folder(REFS)
    groups = build_groups(refs)
    groups = annotate(groups, refs)
    target_fa = build_multi_fasta(groups, refs, tmp)
    results = []
    for name, r1, r2 in [("B4", FASTQS/"B4_R1.fastq.gz", FASTQS/"B4_R2.fastq.gz"),
                          ("G5", FASTQS/"G5_R1.fastq.gz", FASTQS/"G5_R2.fastq.gz")]:
        well_dir = tmp / name
        well_dir.mkdir()
        stats = run_bwa(r1, r2, target_fa, well_dir)
        result = classify(stats.bam_path, groups, refs)
        result.well_id = name
        result = evaluate(result, stats.bam_path, groups, refs)
        results.append(result)
    return results, tmp

def test_json_roundtrip(full_results):
    results, tmp = full_results
    out = tmp / "results.json"
    write_json(results, out)
    assert out.exists()
    loaded = load_json(out)
    assert len(loaded) == 2
    assert loaded[0]["well_id"] == "B4"

def test_well_report_written(full_results):
    results, tmp = full_results
    out = tmp / "report.txt"
    write_report(results, out)
    assert out.exists()
    text = out.read_text()
    assert "B4" in text
    assert "G5" in text

def test_xlsx_written(full_results):
    results, tmp = full_results
    out = tmp / "plate.xlsx"
    write_xlsx(results, out)
    assert out.exists()
    assert out.stat().st_size > 1000  # non-empty workbook

def test_xlsx_has_correct_sheets(full_results):
    results, tmp = full_results
    out = tmp / "plate2.xlsx"
    write_xlsx(results, out)
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert "Plate Map" in wb.sheetnames
    assert "Picks" in wb.sheetnames
