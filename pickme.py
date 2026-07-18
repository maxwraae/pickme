#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pickme.py  —  "Can I pick this clone?" Standalone construct verification.

WHAT YOU DO
-----------
Put this file in a folder that contains:
  * paired-end FASTQs   (*_R1*.fastq.gz  +  *_R2*.fastq.gz)
  * reference maps      (*.gb GenBank files; *.dna SnapGene optional)
...then run:

    python pickme.py          # (Windows: py pickme.py)

It finds the FASTQs and the .gb maps by itself and writes a colored Excel
plate map telling you, for every well: which construct is in it, and whether
it matches the map you ordered.

WORKS FOR ANYONE
----------------
The only thing needed is Python 3.10+. On first run the script pip-installs its
two dependencies (openpyxl + sequence-align), both shipping cross-platform
wheels. No bwa, no samtools, no conda, no compiler — it installs and runs the
same on macOS (Intel & Apple Silicon), Linux, and Windows. It's a pure-Python
reimplementation of the standard "competitive read mapping + pileup variant
calling" workflow, so it reproduces bwa/minimap2 + samtools without them.

HOW IT WORKS (two layers: identity, then quality)
-------------------------------------------------
SETUP   index every reference's 15-mers; group near-identical siblings into
        families; with a real global aligner (sequence-align) map each
        construct's differing positions (used later to flag cross-contamination).
IDENTITY  two stages, the way you'd reason about it by hand:
        1) MATCHES BEST — score every map by how COMPLETELY the reads reconstruct
           it (fraction of its 15-mers seen). A short shared cassette or an
           overlapping parent plasmid can never top a construct the reads fully
           cover, so nothing hijacks the call on shared backbone. Keep the tied
           top maps, and prefer the most complete one (a superset beats its
           subset — e.g. with-intron over no-intron).
        2) UNIQUENESS DISTINGUISHES — when the top maps are near-identical (differ
           only by a small unique region, e.g. a 7-bp barcode), the k-mers unique
           to each decide the exact clone. The winner must DOMINATE the runner-up
           (>=3x) or the distinguishing region wasn't read cleanly -> UNKNOWN.
           A second unique signature present in force = MIX (contamination).
        A well is UNKNOWN when the reads reconstruct no map (<50% of the best),
        or cover the backbone but never span the region that names the construct.
QUALITY  pile reads onto the called construct (seed-chaining handles indels of
        any size). At each position, a non-reference base is a REAL variant only
        if its read count exceeds sequencing error at that depth (binomial test,
        --error-rate). Real variants sort by fraction: >=70% = mutation,
        30-70% = heterogeneous (not a pure clone), <30% = minor subpopulation.
        Differences recurring across most wells of a construct = stale-map
        error, demoted; only well-specific events condemn a clone.
VERDICT  NO_DATA -> RED_UNKNOWN -> RED_MIX -> RED_BROKEN -> RED_MUT -> RED_HET
        -> YELLOW -> GREEN.  Low breadth on an identified well is kept as YELLOW,
        never thrown out.

Output lands in ./pickme_results/ :  plate_map.xlsx  +  .txt  +  .json
See README.md for the full description. Run `python pickme.py --help` for options.

Author: Max Wraae.  MIT License.
"""

from __future__ import annotations

import argparse
import gzip
import html as _html
import json
import math
import multiprocessing as mp
import os
import re
import subprocess
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from pathlib import Path


# ─────────────────────────────────────────────────────────────────────────────
# 0.  Dependency bootstrap — make "just run it" true (openpyxl is pure-Python)
# ─────────────────────────────────────────────────────────────────────────────

def _ensure_deps() -> None:
    need = []
    for mod, pkg in (("openpyxl", "openpyxl"), ("sequence_align", "sequence-align")):
        try:
            __import__(mod)
        except ImportError:
            need.append(pkg)
    if not need:
        return
    print(f"[setup] Installing {', '.join(need)} (one-time) ...", flush=True)
    cmd = [sys.executable, "-m", "pip", "install", "--quiet", *need]
    for extra in ([], ["--user"]):
        try:
            subprocess.check_call(cmd + extra)
            print("[setup] Done.\n", flush=True)
            return
        except subprocess.CalledProcessError:
            continue
    print(f"\nERROR: could not auto-install {', '.join(need)}. Run once yourself:\n"
          f"    {sys.executable} -m pip install {' '.join(need)}\n", file=sys.stderr)
    sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
# 1.  Data model
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class Feature:
    kind: str
    label: str
    start: int
    end: int


@dataclass
class Reference:
    name: str
    sequence: str
    features: list = field(default_factory=list)
    group: int = -1
    aliases: list = field(default_factory=list)  # other map files with identical DNA


@dataclass
class Diff:
    pos: int          # native 0-based reference position
    expected: str
    observed: str
    depth: int
    fraction: float
    kind: str         # "mutation" | "heterogeneous" | "systematic"
    features: list = field(default_factory=list)
    note: str = ""    # e.g. "matches sibling AOPE_50_5"
    ref_count: int = 0
    alt_count: int = 0
    alt_fwd: int = 0  # alt reads on the forward strand ) a real variant is balanced;
    alt_rev: int = 0  # alt reads on the reverse strand ) one-sided = prep/oxidation
    ctx_ref: str = ""     # map sequence +/-15 bp around the change (read in context)
    ctx_obs: str = ""     # what the reads actually said over the same window
    ctx_i: int = 0        # index of the changed base within the context window
    ctx_start: int = 0    # native 0-based position of the first base in the window
    impact: tuple | None = None   # ('silent'|'missense'|'stop'|..., detail) or None
    indel_kind: str = ""  # "ins" | "del" | "" — a small indel, not a substitution
    indel_size: int = 0   # bp inserted (ins) or deleted (del) in the clone
    indel_seq: str = ""   # the inserted bases (for "ins"); "" for "del"


@dataclass
class WellResult:
    sample: str
    plate: str
    well: str
    total_reads: int = 0
    mapped_reads: int = 0
    called: str | None = None
    called_fraction: float = 0.0
    runner_up: str | None = None
    runner_up_fraction: float = 0.0
    breadth: float = 0.0
    best_breadth: float = 0.0    # completeness of the best-matching map (for UNKNOWN)
    identified: bool = False     # did the two-stage engine name a construct?
    is_parent: bool = False      # the named construct is the unmodified parent (empty vector)
    candidates: list = field(default_factory=list)  # top (name, breadth) matches
    # [(name, depth, maxdepth)] — constructs whose sequence IS in this well but which the
    # plate says are background (deep elsewhere, shallow here). Reported as evidence, not
    # as a verdict: the reader sees what is in the well and to what degree, and decides.
    background: list = field(default_factory=list)
    cov_track: list = field(default_factory=list)   # per-bin coverage (0-100%) along the map
    track_len: int = 0           # length of the map the track is drawn against
    track_label: str = ""        # which map the coverage track is against
    track_features: list = field(default_factory=list)  # [start%, end%, label, critical]
    # Per-position read depth against the called map (len == track_len), split by strand:
    # how many reads mapped to each nucleotide from the + vs - strand. The raw material for
    # the radial coverage view — fwd (green) and rev (purple), total = fwd+rev. Empty when
    # no map was called at all.
    depth_fwd: list = field(default_factory=list)
    depth_rev: list = field(default_factory=list)
    mean_depth: float = 0.0
    distinctive_cov: float = 1.0  # fraction of the called construct's distinctive
    has_distinctive: bool = False # positions actually covered by reads
    member_support: int = 0      # reads that uniquely fingerprint the called clone
    mix_flag: bool = False
    n_mut: int = 0               # clone-specific homozygous mutations (disqualifying)
    n_het: int = 0               # heterogeneous positions (two substantial alleles)
    n_minor: int = 0             # real but low-level subpopulation positions
    n_artifact: int = 0          # strand-biased 8-oxoG / prep artifacts (not disqualifying)
    n_systematic: int = 0        # differences shared across the plate = map error
    differences: list = field(default_factory=list)
    deletion: tuple | None = None      # (start, end, bp) suspected internal deletion
    deletion_features: list = field(default_factory=list)
    uncovered: list = field(default_factory=list)  # [(start,end)] runs with no reads
    called_idx: int = -1               # index into refs of the called map
    corrected_gb: str = ""             # downloadable read-consensus GenBank (per well)
    verdict: str = "NO_DATA"
    reason: str = ""
    # 1 byte per position of the called map: was there enough depth to judge it at all?
    # A well that never covered a position must not count as evidence AGAINST a map
    # error there (see mark_systematic). In-memory only — stripped by write_json.
    assayed: bytes = b""


# ─────────────────────────────────────────────────────────────────────────────
# 2.  Sequence helpers
# ─────────────────────────────────────────────────────────────────────────────

_COMP = str.maketrans("ACGTNacgtn", "TGCANtgcan")


def revcomp(s: str) -> str:
    return s.translate(_COMP)[::-1]


# standard genetic code (stop = *), for the "does this base change matter?" call
_CODON = {
    'TTT': 'F', 'TTC': 'F', 'TTA': 'L', 'TTG': 'L', 'CTT': 'L', 'CTC': 'L',
    'CTA': 'L', 'CTG': 'L', 'ATT': 'I', 'ATC': 'I', 'ATA': 'I', 'ATG': 'M',
    'GTT': 'V', 'GTC': 'V', 'GTA': 'V', 'GTG': 'V', 'TCT': 'S', 'TCC': 'S',
    'TCA': 'S', 'TCG': 'S', 'CCT': 'P', 'CCC': 'P', 'CCA': 'P', 'CCG': 'P',
    'ACT': 'T', 'ACC': 'T', 'ACA': 'T', 'ACG': 'T', 'GCT': 'A', 'GCC': 'A',
    'GCA': 'A', 'GCG': 'A', 'TAT': 'Y', 'TAC': 'Y', 'TAA': '*', 'TAG': '*',
    'CAT': 'H', 'CAC': 'H', 'CAA': 'Q', 'CAG': 'Q', 'AAT': 'N', 'AAC': 'N',
    'AAA': 'K', 'AAG': 'K', 'GAT': 'D', 'GAC': 'D', 'GAA': 'E', 'GAG': 'E',
    'TGT': 'C', 'TGC': 'C', 'TGA': '*', 'TGG': 'W', 'CGT': 'R', 'CGC': 'R',
    'CGA': 'R', 'CGG': 'R', 'AGT': 'S', 'AGC': 'S', 'AGA': 'R', 'AGG': 'R',
    'GGT': 'G', 'GGC': 'G', 'GGA': 'G', 'GGG': 'G',
}
_AA3 = {'A': 'Ala', 'R': 'Arg', 'N': 'Asn', 'D': 'Asp', 'C': 'Cys', 'Q': 'Gln',
        'E': 'Glu', 'G': 'Gly', 'H': 'His', 'I': 'Ile', 'L': 'Leu', 'K': 'Lys',
        'M': 'Met', 'F': 'Phe', 'P': 'Pro', 'S': 'Ser', 'T': 'Thr', 'W': 'Trp',
        'Y': 'Tyr', 'V': 'Val', '*': 'Stop', 'X': '?'}


# features worth translating: exons / ORFs / CDS / genes (not backbone origins etc.)
_CODING_RE = re.compile(r"exon|cds|\borf\b|gene|payload|mecp2|transgene", re.I)


def _translate(seq: str) -> str:
    return "".join(_CODON.get(seq[i:i + 3], "X")
                   for i in range(0, len(seq) - 2, 3))


def _codon_effect(codon, obs_codon):
    """(label, detail) for a single ref->observed codon substitution."""
    aa_r, aa_o = _CODON.get(codon, "X"), _CODON.get(obs_codon, "X")
    if aa_o == aa_r:
        return ("silent", f"{_AA3.get(aa_r)} unchanged — synonymous")
    if aa_o == "*":
        return ("stop", f"premature STOP ({_AA3.get(aa_r)}→Stop) — truncates protein")
    if aa_r == "*":
        return ("readthrough", f"Stop→{_AA3.get(aa_o)} — read-through")
    return ("missense", f"{_AA3.get(aa_r)}→{_AA3.get(aa_o)}")


def _coding_impact(ref_seq, pos, observed, feat_start, feat_end):
    """Predict whether a base change matters at the PROTEIN level, for a change inside a
    coding-ish feature (an exon / CDS). We don't store the reading frame, so we infer it
    HONESTLY: of the six frames (3 per strand), keep those that translate the WHOLE
    feature with NO internal stop codon — a real ORF has exactly one such frame. We then
    assert only an effect the surviving frame(s) actually agree on:
      * exactly one ORF frame            -> a confident silent/missense/stop/read-through call;
      * several frames, all in agreement  -> that same call;
      * frames disagree, or NONE is a clean ORF (a short or internal exon whose frame is
        set by the upstream exons we don't have) -> ('uncertain', ...), stated plainly
        rather than inventing an amino-acid change.
    Returns (label, detail), or None when there is no codon to read.  Display-only: this
    never changes a variant's kind or the well's verdict."""
    seg = ref_seq[feat_start:feat_end]
    if len(seg) < 6:
        return None
    # plausible reading frames = zero INTERNAL stop codons across the whole feature
    # (a single trailing stop, i.e. the CDS's own terminator, is allowed and ignored).
    plausible = []
    for strand in ("+", "-"):
        s = seg if strand == "+" else revcomp(seg)
        for frame in (0, 1, 2):
            tr = _translate(s[frame:])
            internal = tr[:-1].count("*") if tr.endswith("*") else tr.count("*")
            if internal == 0:
                plausible.append((strand, frame))
    UNCERTAIN_NOFRAME = ("uncertain",
        "protein effect undetermined — no open reading frame spans this feature "
        "(likely a short or internal exon whose frame is set by the upstream exons)")
    if not plausible:
        return UNCERTAIN_NOFRAME
    outcomes = []
    for strand, frame in plausible:
        s = seg if strand == "+" else revcomp(seg)
        if strand == "+":
            p, ob = pos - feat_start, observed
        else:
            p, ob = (feat_end - 1) - pos, revcomp(observed)
        rel = p - frame
        if rel < 0:
            continue
        cstart = frame + (rel // 3) * 3
        if cstart + 3 > len(s):
            continue
        within = p - cstart
        codon = s[cstart:cstart + 3]
        obs_codon = codon[:within] + ob + codon[within + 1:]
        if any(b not in "ACGT" for b in codon + obs_codon):
            continue
        outcomes.append(_codon_effect(codon, obs_codon))
    if not outcomes:
        return None
    labels = {o[0] for o in outcomes}
    details = {o[1] for o in outcomes}
    if len(details) == 1:                         # every frame agrees exactly
        return outcomes[0]
    if len(labels) == 1:                          # same class, residues differ (rare)
        return (outcomes[0][0],
                outcomes[0][1] + " — reading frame ambiguous, residues approximate")
    return ("uncertain",
            "protein effect ambiguous — this feature has multiple possible reading frames "
            f"giving different results ({'; '.join(sorted(d for _, d in outcomes))})")


def _ctx_window(ref_seq, base_counts, depth, pos, alt, flank=15):
    """The local sequence at a change: what the MAP says vs what the READS say,
    +/-flank bases, so the change can be read in context (splice motif? start
    codon?). Neighbours = consensus where covered, '.' where not sequenced. The
    CHANGED base itself is forced to the called allele `alt` (for a minor the
    consensus is still the reference, but we want to show the variant we describe)."""
    lo, hi = max(0, pos - flank), min(len(ref_seq), pos + flank + 1)
    rmap = ref_seq[lo:hi]
    robs = []
    for p in range(lo, hi):
        if p == pos:
            robs.append(alt)
        elif depth[p] <= 0:
            robs.append(".")
        else:
            robs.append(max("ACGT", key=lambda b: base_counts[p][b]))
    return rmap, "".join(robs), pos - lo    # ref, obs, index of the changed base


# ─────────────────────────────────────────────────────────────────────────────
# 3.  Reference loading  (pure-Python GenBank parser — no Biopython needed)
# ─────────────────────────────────────────────────────────────────────────────

def parse_genbank(path: Path) -> Reference:
    text = path.read_text(errors="replace")
    seq_chars, feature_lines = [], []
    in_origin = in_features = False
    for line in text.splitlines():
        if line.startswith("ORIGIN"):
            in_origin, in_features = True, False
            continue
        if line.startswith("FEATURES"):
            in_features = True
            continue
        if in_origin:
            if line.startswith("//"):
                in_origin = False
                continue
            seq_chars.append(re.sub(r"[^A-Za-z]", "", line))
        elif in_features:
            if line and not line.startswith(" "):
                in_features = False
            else:
                feature_lines.append(line)
    sequence = "".join(seq_chars).upper()

    features: list[Feature] = []
    cur_kind = cur_label = cur_span = None
    loc_re = re.compile(r"(\d+)\.\.(\d+)")

    def flush():
        if cur_kind and cur_span and cur_kind != "source":
            features.append(Feature(cur_kind, cur_label or cur_kind,
                                    cur_span[0], cur_span[1]))

    for line in feature_lines:
        m = re.match(r"\s{5}(\S+)\s+(.*)", line)
        if m and not line.lstrip().startswith("/"):
            flush()
            cur_kind, cur_label = m.group(1), None
            nums = loc_re.findall(m.group(2))
            if nums:
                cur_span = (min(int(a) for a, _ in nums) - 1,
                            max(int(b) for _, b in nums))
            else:
                cur_span = None
        else:
            qm = re.search(r'/(?:label|gene|product|note)="?([^"]+)"?', line)
            if qm and cur_label is None:
                cur_label = qm.group(1).strip()
    flush()
    return Reference(name=path.stem, sequence=sequence, features=features)


def try_parse_snapgene(path: Path) -> Reference | None:
    # Prefer the full library if the user happens to have it (it also gives
    # features), but never depend on it — fall back to a built-in reader below.
    try:
        import snapgene_reader  # type: ignore
        data = snapgene_reader.parse(str(path))
        seq = (data.get("seq") or data.get("sequence") or "").upper()
        feats = [Feature(f.get("type", "misc_feature"),
                         f.get("label", f.get("name", "feature")),
                         int(f.get("start", 0)), int(f.get("end", 0)))
                 for f in (data.get("features", []) or [])]
        if seq:
            return Reference(path.stem, seq, feats)
    except Exception:
        pass
    return _read_snapgene_native(path)


def _read_snapgene_native(path: Path) -> Reference | None:
    """Dependency-free SnapGene (.dna) sequence reader.

    A .dna file is a flat series of blocks: [1-byte type][4-byte big-endian
    length][payload].  Block type 0 is the DNA sequence — payload[0] is a
    topology/flags byte, payload[1:] is the bases.  (Feature blocks are XML;
    we skip them here — pickme only needs the sequence to call a clone.)
    """
    try:
        raw = path.read_bytes()
    except Exception:
        return None
    i, n, seq = 0, len(raw), None
    while i + 5 <= n:
        seg_type = raw[i]
        seg_len = int.from_bytes(raw[i + 1:i + 5], "big")
        i += 5
        if seg_len < 0 or i + seg_len > n:
            break
        payload = raw[i:i + seg_len]
        i += seg_len
        if seg_type == 0 and len(payload) > 1:
            seq = payload[1:].decode("ascii", "ignore").upper()
            break
    if seq and re.fullmatch(r"[ACGTN]+", seq):
        return Reference(path.stem, seq, [])
    return None


def load_references(refs: list[Path]) -> list[Reference]:
    out: list[Reference] = []
    for p in sorted(refs):
        if p.suffix.lower() == ".gb":
            try:
                r = parse_genbank(p)
                if r.sequence:
                    out.append(r)
                else:
                    print(f"  ! {p.name}: no sequence, skipped")
            except Exception as e:
                print(f"  ! {p.name}: parse failed ({e}), skipped")
        elif p.suffix.lower() == ".dna":
            r = try_parse_snapgene(p)
            if r:
                out.append(r)
            else:
                print(f"  ! {p.name}: could not read SnapGene sequence, skipped")
    seen = {}
    for r in out:
        seen.setdefault(r.name, r)
    by_name = list(seen.values())
    # Collapse map files whose DNA is byte-identical (same sequence, different
    # name/annotation). No read can tell them apart, so treating them as separate
    # references would leave every matching well falsely "ambiguous". Keep one,
    # record the others as aliases, and say so out loud.
    by_seq: dict[str, Reference] = {}
    for r in by_name:
        rep = by_seq.get(r.sequence)
        if rep is None:
            by_seq[r.sequence] = r
        else:
            rep.aliases.append(r.name)
    for r in by_seq.values():
        if r.aliases:
            print(f"  note: {r.name} has IDENTICAL DNA to {', '.join(r.aliases)} "
                  f"— treated as one map (check your exports)")
    return list(by_seq.values())


# ─────────────────────────────────────────────────────────────────────────────
# 4.  k-mer index
# ─────────────────────────────────────────────────────────────────────────────

K = 15


def build_index(refs: list[Reference]):
    """kmer -> list of (ref_idx, strand, forward_strand_pos)."""
    index: dict[str, list] = defaultdict(list)
    for ri, ref in enumerate(refs):
        seq, n = ref.sequence, len(ref.sequence)
        for i in range(n - K + 1):
            index[seq[i:i + K]].append((ri, 1, i))
        rc = revcomp(seq)
        for i in range(n - K + 1):
            index[rc[i:i + K]].append((ri, -1, n - i - K))
    return index


# ─────────────────────────────────────────────────────────────────────────────
# 5.  Seed-chaining pileup  (the engine)
# ─────────────────────────────────────────────────────────────────────────────

def seedchain_tally(read: str, qual, target_idx: int, ref_len: int,
                    index, depth: list, base_counts, strand_counts,
                    indel_counts=None):
    """Place `read` onto reference `target_idx` by chaining its matching seed
    blocks, and tally per-position base counts. A jump between colinear blocks
    is an indel (handled, not smeared). Returns True if the read was placed.

    When two solid seed blocks chain on DIFFERENT diagonals, the read carries a
    small indel vs the map (a frameshift-class event a substitution pileup would
    miss or mislabel). We record it into `indel_counts` so the caller can call it
    with the same error-floor + both-strand test used for substitutions.

    Two prep-artifact defenses baked in here:
      * a base whose Phred quality is below MIN_BASE_QUAL is dropped (not counted
        toward depth or the base tally) — low-quality calls are the main error
        source;
      * strand_counts records, per position and base, how many reads came from
        the + vs - strand, so a strand-skewed call (8-oxoG G>T / C>A damage, which
        hits one strand only) can be told apart from a real both-strand variant."""
    # gather seeds for this target on both strands
    fwd, rev = [], []
    L = len(read)
    for off in range(0, L - K + 1):
        for (ri, st, pos) in index.get(read[off:off + K], ()):
            if ri != target_idx:
                continue
            (fwd if st == 1 else rev).append((off, pos))
    if len(fwd) + len(rev) < 3:
        return False
    use_rev = len(rev) > len(fwd)
    seeds = rev if use_rev else fwd
    oriented = revcomp(read) if use_rev else read
    # quality travels with the read; reversing the read reverses the qual string
    oq = (qual[::-1] if use_rev else qual) if qual else None
    have_q = oq is not None and len(oq) == L
    si = 1 if use_rev else 0        # strand slot: forward-strand reads -> 0, reverse -> 1
    # convert to oriented-read offset + diagonal (ref_pos - oriented_off)
    pts = []
    for off, pos in seeds:
        ooff = (L - K - off) if use_rev else off
        pts.append((ooff, pos - ooff))
    pts.sort()
    # group seeds into blocks of constant diagonal, split when offsets jump >40
    bydiag = defaultdict(list)
    for ooff, diag in pts:
        bydiag[diag].append(ooff)
    blocks = []
    for diag, offs in bydiag.items():
        offs.sort()
        s = p = offs[0]
        for o in offs[1:]:
            if o - p > 40:
                blocks.append((s, p + K, diag))
                s = o
            p = o
        blocks.append((s, p + K, diag))
    # tally; assign each oriented position once, largest blocks win overlaps
    covered = bytearray(L)
    for (omin, omax, diag) in sorted(blocks, key=lambda b: -(b[1] - b[0])):
        for o in range(max(0, omin), min(L, omax)):
            if covered[o]:
                continue
            covered[o] = 1
            a = oriented[o]
            if a not in "ACGT":
                continue
            if have_q and (ord(oq[o]) - 33) < MIN_BASE_QUAL:
                continue                        # drop a low-quality base call
            rp = (o + diag) % ref_len
            depth[rp] += 1
            base_counts[rp][a] += 1
            sc = strand_counts[rp].get(a)
            if sc is None:
                strand_counts[rp][a] = [0, 0]
                sc = strand_counts[rp][a]
            sc[si] += 1
    # ── small-indel junctions: two solid blocks on different diagonals ──
    # A single clean alignment is one diagonal; a jump to a neighbouring diagonal
    # over a short read gap is a small insertion/deletion in the clone. Require
    # BOTH flanking blocks to be solid (>= K oriented bases) and the read gap to
    # match the diagonal shift, so repeats and mapping wobble don't fabricate one.
    if indel_counts is not None:
        ordered = sorted(blocks, key=lambda b: b[0])
        seen_j = set()
        for (a_min, a_max, a_diag), (b_min, b_max, b_diag) in zip(ordered, ordered[1:]):
            if a_diag == b_diag:
                continue
            if (a_max - a_min) < K or (b_max - b_min) < K:
                continue                          # a flanking block is too thin to trust
            d = abs(a_diag - b_diag)
            if not (1 <= d <= 30):
                continue                          # only SMALL indels here (big = deletion dip)
            read_gap = b_min - a_max
            ins = b_diag < a_diag                 # read runs ahead of ref -> extra bases = insertion
            # read_gap should be ~+d for an insertion, ~0 for a deletion; allow slack
            if ins and abs(read_gap - d) > 3:
                continue
            if not ins and abs(read_gap) > 3:
                continue
            rj = (a_max + a_diag) % ref_len       # ref coord just past block A = the junction
            kind = "ins" if ins else "del"
            key = (rj, kind, d)
            if key in seen_j:
                continue                          # one vote per read per junction
            seen_j.add(key)
            iseq = ""
            if ins:
                iseq = oriented[a_max:a_max + d]
                if any(c not in "ACGT" for c in iseq) or len(iseq) != d:
                    continue
            slot = indel_counts.get(key)
            if slot is None:
                slot = indel_counts[key] = [0, 0, Counter()]
            slot[si] += 1
            if ins:
                slot[2][iseq] += 1
    return True


def _hirschberg_pair(task):
    """Worker: global-align one ordered (sibling, target) pair. Returns the pair key
    plus the aligned sequences so the parent can rebuild the cache in any order."""
    from sequence_align.pairwise import hirschberg
    qi, ti, seq_q, seq_t = task
    qa, qb = hirschberg(list(seq_q), list(seq_t), gap="-",
                        match_score=1, mismatch_score=-1, indel_score=-1)
    return qi, ti, qa, qb


def build_distinctive(refs, ref_by_group, jobs=1):
    """Once, with a REAL global aligner: for every construct, find the positions
    that distinguish it from its same-family siblings (the insert / variant
    region). Seed-chaining can't do this — it skips exactly the differing
    positions — so we use sequence-align's Needleman-Wunsch (Hirschberg).
    Returns {ref_idx: {target_pos: {sibling_idx: base}}}.

    The pairwise alignments are O(family²) and dominate setup, so with jobs>1 we fan
    them across a spawn Pool. The result is order-independent (deterministic aligner +
    fixed assembly), so parallel output is identical to the jobs==1 serial path."""
    from sequence_align.pairwise import hirschberg

    # every ordered (sibling m, target ci) pair we need, deduplicated
    needed, seen = [], set()
    for members in ref_by_group.values():
        for ci in members:
            for m in members:
                if m != ci and (m, ci) not in seen:
                    seen.add((m, ci))
                    needed.append((m, ci))

    pair_cache = {}
    if jobs and jobs != 1 and len(needed) > 1:
        ctx = mp.get_context("spawn")
        tasks = [(qi, ti, refs[qi].sequence, refs[ti].sequence) for qi, ti in needed]
        with ctx.Pool(processes=min(jobs, len(tasks))) as pool:
            for qi, ti, qa, qb in pool.imap_unordered(_hirschberg_pair, tasks):
                pair_cache[(qi, ti)] = (qa, qb)
    else:
        for qi, ti in needed:
            pair_cache[(qi, ti)] = hirschberg(
                list(refs[qi].sequence), list(refs[ti].sequence), gap="-",
                match_score=1, mismatch_score=-1, indel_score=-1)

    out = {}
    for members in ref_by_group.values():
        for ci in members:
            target = refs[ci].sequence
            disc = {}
            for m in members:
                if m == ci:
                    continue
                qa, qb = pair_cache[(m, ci)]  # qa=sibling, qb=target(called)
                tp = 0
                for x, y in zip(qa, qb):
                    if y != "-":             # consumes a target position
                        if x != "-" and x != target[tp] and x in "ACGT":
                            disc.setdefault(tp, {})[m] = x
                        tp += 1
            out[ci] = disc
    return out


# ─────────────────────────────────────────────────────────────────────────────
# 6.  FASTQ reading
# ─────────────────────────────────────────────────────────────────────────────

def read_fastq_seqs(path: Path, cap: int | None = None):
    """Yield (sequence, quality) per read. Quality (Phred+33) is kept so the
    pileup can drop low-quality base calls — the main defense against prep noise."""
    op = gzip.open if path.suffix == ".gz" else open
    n = 0
    with op(path, "rt") as fh:
        while True:
            h = fh.readline()
            if not h:
                break
            seq = fh.readline().strip().upper()
            fh.readline()               # "+" separator
            qual = fh.readline().rstrip("\n")
            if seq:
                yield seq, qual
                n += 1
                if cap and n >= cap:
                    break


# ─────────────────────────────────────────────────────────────────────────────
# 7.  Reference grouping (k-mer Jaccard -> union-find)
# ─────────────────────────────────────────────────────────────────────────────

def group_refs(refs: list[Reference]):
    sets = []
    for r in refs:
        s = set()
        for i in range(0, len(r.sequence) - K + 1, 7):
            s.add(r.sequence[i:i + K])
        sets.append(s)
    parent = list(range(len(refs)))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    for i in range(len(refs)):
        for j in range(i + 1, len(refs)):
            if sets[i] and sets[j]:
                if len(sets[i] & sets[j]) / min(len(sets[i]), len(sets[j])) >= 0.5:
                    parent[find(j)] = find(i)
    groups = defaultdict(list)
    for i, r in enumerate(refs):
        r.group = find(i)
        groups[r.group].append(i)
    return groups


# ─────────────────────────────────────────────────────────────────────────────
# 8.  Per-well evaluation
# ─────────────────────────────────────────────────────────────────────────────

MIN_MAPPED = 100
# ── Identity (two stages) ────────────────────────────────────────────────────
# Stage 1 — "what does it match best": rank references by how COMPLETELY the
# reads reconstruct them (fraction of the map's k-mers seen). A short shared
# cassette or a high-depth sub-region can never top a fully-covered construct,
# so this alone stops a parent/overlap map from hijacking the call.
TIER_MARGIN = 0.08            # references within this breadth of the best are "tied"
MIN_IDENTIFY_BREADTH = 0.50   # need at least half a map reconstructed to call it; below
                              # this the reads are too sparse to trust an identity -> UNKNOWN
                              # (an identified well with 50-90% breadth is kept, flagged YELLOW)
# Stage 2 — "uniqueness distinguishes": among the tied top maps, the k-mers that
# belong to exactly ONE of them (the distinctive region, e.g. the barcode) pick
# the exact construct; a second such signature present = contamination.
MIN_MEMBER_SUPPORT = 30       # reads carrying a construct's unique k-mers to call it.
                              # A well that lacks a barcode still accrues ~5-20 stray
                              # unique-k-mer reads (index hopping, low-complexity/A-rich
                              # barcodes, sequencing error); genuinely-present barcodes
                              # draw 45+. 30 sits in the empty gap between the two.
DOMINANCE = 3.0               # winner must beat the runner-up by this ratio, else the
                              # distinguishing region is too weak/ambiguous to call
MIX_UNIQUE_FRACTION = 0.20    # a 2nd construct with >=20% of the winner's unique reads = mix
CLOSE_MARGIN = 0.02           # maps within this breadth of the best are "near-ties" that
                              # must be separated by their unique k-mers (barcode); a map
                              # covered clearly less well than the best simply loses
# A fingerprint k-mer must be real sequence, not a homopolymer/short-repeat: those are
# shared by every poly-N run in the reads (Illumina slippage extends them), so a map
# whose only "unique" k-mers are low-complexity would collect votes from EVERY well — a
# magnet that drowns the true signal. Siblings that differ ONLY by a run's LENGTH (a
# poly-T tract of 12 vs 14 vs 17 bp) are instead separated by measuring that length
# directly, anchored on the specific sequence flanking the run.
LOWCOMPLEX_MAXBASE = 0.6      # a k-mer with one base >60% of its length is homopolymer-ish
LOWCOMPLEX_MIN_DINUC = 6      # ...or built from <6 distinct dinucleotides = a short repeat
RUN_MIN = 6                   # a homopolymer run this long can be a length-distinguishing locus
RUN_ANCHOR = 12              # bp of specific (non-low-complexity) flank that pins the run's end
SIG_CALL_FRAC = 0.70          # fraction of a map's OWN unique k-mers the reads must carry
                              # before the well may be called that map. Fit says "these
                              # reads are consistent with X"; this says "X is actually
                              # here". Without it a plate-wide contaminant wins every well
                              # whose real construct is unreadable.
RUN_SLIP = 3                  # bp tolerance for polymerase slippage / off-target runs. Illumina
                              # almost always UNDER-counts a homopolymer, so a read showing L bases
                              # is credited to the shortest construct whose tract is >= L (reads
                              # longer than every candidate fall to the longest). This stops a true
                              # long tract that slipped short from being mis-read as a shorter
                              # sibling — the false-MIX / mis-call failure mode on poly-N families.
# ── Quality layer ────────────────────────────────────────────────────────────
BREADTH_GREEN = 0.90          # below this a called well is kept but flagged YELLOW
DIFF_MIN_DEPTH = 10           # need this depth to judge a position at all
DIFF_MIN_READS = 3            # min reads of a sibling base to count it as present
SIBLING_MIX_FRACTION = 0.15   # a 15% sibling contaminant at a diff position flags a mix
# A base is a REAL variant only if its read count exceeds what sequencing error
# explains at that depth (a binomial test), not a flat %.
ERROR_RATE = 0.01             # assumed per-base sequencing error (set with --error-rate)
Z_SIGNIF = 5.0                # ~genome-wide significance for the error test
MIN_ALT_READS = 3             # absolute floor before any test
# once a base is statistically REAL, its fraction sorts it:
MUT_FRACTION = 0.70           # alt this dominant -> homozygous mutation
HET_FRACTION = 0.30           # both alleles this substantial -> heterogeneous (not pure)
# Prep-artifact defenses. A base below this Phred quality is not counted at all; a
# "real" variant that shows on essentially one DNA strand only is 8-oxoG oxidation
# (G>T / C>A) or a library artifact, not a clone mutation, so it is demoted.
MIN_BASE_QUAL = 20            # drop bases with Phred < 20 (>1% error) before counting
STRAND_MIN_ALT = 6            # only apply the strand test once the alt has this many reads
STRAND_MIN_FRAC = 0.10        # a real variant has >=10% of its reads on the minority strand
THIN_DEPTH = 15
DELETION_MIN_BP = 30
SYSTEMATIC_FRACTION = 0.50    # a fixed diff in >=half a family's wells = map error
COORD_FAMILY_MAXDIFF = 0.02   # same-length maps differing at <=2% of positions are
                              # barcode siblings off one parent: one coordinate frame,
                              # so they share map errors at identical positions


def _read_kmers(reads):
    """Every k-mer present in the reads, both orientations, deduplicated. Cheap
    membership set for scoring how completely the reads cover each reference."""
    ks = set()
    for rd in reads:
        for s in (rd, revcomp(rd)):
            for o in range(len(s) - K + 1):
                ks.add(s[o:o + K])
    return ks


def _coverage(refs, readset):
    """Per reference: (idx, covered_kmers, breadth) sorted by BREADTH (best fit
    first). breadth = fraction of the map's own k-mers seen in the reads = how
    completely the reads reconstruct THIS map. A short cassette or a parent that
    the reads only partly match can never top a construct they fully cover."""
    out = []
    for ri, ref in enumerate(refs):
        m = len(ref.sequence) - K + 1
        if m <= 0:
            continue
        seq = ref.sequence
        c = sum(1 for i in range(m) if seq[i:i + K] in readset)
        out.append((ri, c, c / m))
    out.sort(key=lambda x: -x[2])
    return out


def _low_complexity(km):
    """True if a k-mer is too repetitive to be a trustworthy fingerprint — one base
    dominates it, or it is built from too few distinct dinucleotides (a homopolymer or
    short repeat). Such k-mers occur inside every poly-N run in the reads, so they must
    not be used to fingerprint a construct (see the LOWCOMPLEX_* rationale)."""
    c = Counter(km)
    if max(c.values()) > LOWCOMPLEX_MAXBASE * len(km):
        return True
    if len({km[i:i + 2] for i in range(len(km) - 1)}) < LOWCOMPLEX_MIN_DINUC:
        return True
    return False


_SIG_CACHE: dict = {}


def _signature_owner(refs):
    """k-mer -> ref_idx for k-mers unique to ONE map across the WHOLE supplied set.

    Uniqueness MUST be judged against every reference, not just the candidates being
    compared. Judged only among candidates, a k-mer counts as "distinctive of X"
    whenever the other candidates happen to lack it — which is evidence that their
    variant is ABSENT, not that X's is PRESENT. The two differ exactly when the reads
    carry something not in the candidate set (an unmodified parent, a failed clone):
    those reads then vote, confidently, for whichever sibling their sequence survives
    in the longest. Scoring against all maps drops such k-mers (some other map has
    them too), so a construct is named only on sequence unique to it.

    Pure function of `refs`, so it is built once per process and reused by every well.
    """
    hit = _SIG_CACHE.get(id(refs))
    if hit is not None and hit[0] is refs:
        return hit[1]

    owner = _tier_unique(range(len(refs)), refs)
    sets = defaultdict(set)
    for km, ri in owner.items():
        sets[ri].add(km)
    _SIG_CACHE[id(refs)] = (refs, owner, dict(sets))  # hold refs so its id() can't recycle
    return owner


def _signature_sets(refs):
    """ref_idx -> its set of map-unique k-mers (see _signature_owner)."""
    _signature_owner(refs)
    return _SIG_CACHE[id(refs)][2]


def _tier_unique(cands, refs):
    """k-mer -> ref_idx for HIGH-complexity k-mers present in exactly ONE of `cands`.
    These are the distinctive signatures (barcode / insert junction / point edit) that
    separate otherwise-identical candidates. Low-complexity k-mers are excluded — they
    are shared by every poly-N run in the reads and would fingerprint nothing reliably
    (siblings differing only by run LENGTH are handled by _length_loci instead)."""
    owner = {}
    for ri in cands:
        seq = refs[ri].sequence
        for km in {seq[i:i + K] for i in range(len(seq) - K + 1)}:
            if _low_complexity(km):
                continue
            owner[km] = ri if km not in owner else None
    return {km: ri for km, ri in owner.items() if ri is not None}


def _length_loci(cands, refs):
    """Homopolymer runs that distinguish the candidates by LENGTH (e.g. a poly-T tract
    present at 12 / 14 / 17 bp across three siblings). k-mers cannot separate these — the
    runs share every sub-window — so we anchor on the specific sequence flanking the run's
    3' end and read the length off directly. Returns
    [(run_base, anchor_kmer, {ref_idx: run_length}), ...] for runs whose length differs
    across cands and whose flank is a specific (non-low-complexity) anchor."""
    tmp = {}
    for ri in cands:
        s = refs[ri].sequence
        i, n = 0, len(s)
        while i < n:
            j = i
            while j < n and s[j] == s[i]:
                j += 1
            if j - i >= RUN_MIN and j + RUN_ANCHOR <= n:
                anchor = s[j:j + RUN_ANCHOR]
                if not _low_complexity((anchor + "AAA")[:K]):
                    tmp.setdefault((s[i], anchor), {})[ri] = j - i
            i = j
    return [(b, a, d) for (b, a), d in tmp.items() if len(set(d.values())) > 1]


def _unique_votes(reads, cands, refs):
    """How many reads carry each candidate's distinguishing signature. Two complementary,
    additive signals (one vote per read per candidate per signal):
      * high-complexity k-mers unique to one candidate (a barcode / insert / point edit);
      * homopolymer run LENGTH at loci where the candidates differ only in run size, read
        off against the specific flank anchor (so a 12-vs-14-vs-17 bp poly-T tract, which
        shares every k-mer, is still resolved)."""
    cs = set(cands)
    tu = {km: ri for km, ri in _signature_owner(refs).items() if ri in cs}
    loci = _length_loci(cands, refs)
    v = Counter()
    for rd in reads:
        strands = (rd, revcomp(rd))
        got = set()
        for s in strands:
            for o in range(len(s) - K + 1):
                ri = tu.get(s[o:o + K])
                if ri is not None:
                    got.add(ri)
        for ri in got:
            v[ri] += 1
        if loci:
            hit = set()
            for s in strands:
                for base, anchor, d in loci:
                    p = s.find(anchor)
                    if p <= 0:                      # anchor absent, or run runs off read start
                        continue
                    q = p
                    while q > 0 and s[q - 1] == base:
                        q -= 1
                    if q == 0:                      # run truncated by the read end -> unknown length
                        continue
                    L = p - q
                    lengths = list(d.values())
                    lo, hi = min(lengths), max(lengths)
                    if L < lo - RUN_SLIP or L > hi + RUN_SLIP:
                        continue                    # a homopolymer unrelated to this locus
                    # slippage under-counts: attribute the read to the SHORTEST construct whose
                    # tract is at least as long as observed (over-long reads fall to the longest).
                    geq = [l for l in lengths if l >= L]
                    target = min(geq) if geq else hi
                    winners = [ri for ri, l in d.items() if l == target]
                    if len(winners) == 1:           # a shared length can't distinguish -> k-mers do
                        hit.add(winners[0])
            for ri in hit:
                v[ri] += 1
    return v


# ── Base-level identity (primary) ────────────────────────────────────────────
# K-mer uniqueness is only a PROXY for "does the well contain this construct's own
# sequence", and the proxy collapses when two maps are near-identical — their unique
# k-mers mask each other out, leaving nothing to identify them by (the Exon family:
# six overlapping maps, 0-25 unique k-mers each). The base-level engine decides on
# ALIGNED BASES instead. For each shortlisted candidate it aligns the reads (_pileup)
# and judges the candidate's DEFINING region — the sequence unique to it among the
# shortlist (its barcode variant / insert / cargo) — by two things:
#   * COVERAGE: is the defining region sequenced, at depth comparable to the well's
#     OWN backbone depth? (self-normalizing, so uneven Tn5 coverage can't fool it)
#   * IDENTITY: do the consensus bases there MATCH this candidate specifically?
# Coverage answers "is the insert physically present"; identity answers "which
# construct". A foreign construct on the shared backbone fails identity; a well
# missing its insert fails coverage. Immune to mutual masking and the length confound.
SHORTLIST_MAX = 16            # cap candidates aligned per well (bounds _pileup cost)
IDENT_READ_CAP = 8000         # pairs sampled for the shortlist alignment pass
MIN_DEFINING_BP = 12          # a candidate needs this many defining positions to be judged
DEF_COV_FRAC = 0.80           # >= this fraction of the defining region must have reads. A
                              # real construct covers ~ALL of its own defining region (cov
                              # ~1.0); a plate-wide background contaminant (pSL0423) smears
                              # across ~0.6, and a near-parent barcode variant matched only
                              # via scattered parental k-mers sits ~0.74 — this floor keeps
                              # the real calls and drops both of those confident-wrong ways
DEF_MIN_SUPPORT = 15          # ...totalling >= this many read-bases (NOT a fraction of the
                              # backbone depth: Tn5 systematically under-covers the barcode
                              # locus here, so a depth-fraction gate wrongly rejects real
                              # thin-but-covered barcodes — an absolute floor is honest)
DEF_ID_MIN = 0.90             # consensus bases must match the candidate at >= this fraction
MIX_FRAC = 0.25               # a 2nd construct is really present only if, at the positions
                              # that distinguish it from the winner, its own bases carry
                              # >= this read fraction (shared flanks don't count — that was
                              # inflating wrong siblings to a false "mix")
PARENT_ID_MIN = 0.97          # empty-vector call: reads must rebuild the parent backbone at
                              # >= this identity over a substantial covered span
PARENT_MIN_COV = 500          # ...and cover at least this many parent bases
SIG_PRESENT_FRAC = 0.60       # unbiased presence: a construct whose OWN map-unique signature
                              # is >= this fraction present in the reads is really in the well,
                              # regardless of its label or of breadth ranking. Catches a clone
                              # that landed in the WRONG well (a pick/plate swap) and inserts
                              # too long to ever win breadth. (The plate contaminant pSL0423
                              # tops out ~0.52, so 0.60 does not fire on background.)
SIG_MIN_SIZE = 20            # ...and the signature must have >= this many k-mers to trust it
CONTAM_MIN_WELLS = 5         # a construct CALLED in >= this many wells is a plate-wide
                             # background candidate (e.g. pSL0423 here). Its shallow calls
CONTAM_DEPTH_FRAC = 0.20     # (depth < this fraction of the deepest well that called it,
CONTAM_ABS_DEPTH = 50        #  OR below this absolute depth) are the contaminant, not the
                             #  well's own clone, and get demoted to "can't tell". A real
                             #  localised construct (an Exon in 1-3 wells) is never touched.


def _median(xs):
    s = sorted(xs)
    n = len(s)
    if not n:
        return 0.0
    return s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2


def _is_parent_map(name):
    """The unmodified backbone map (no barcode / no insert) — its defining region is
    the ABSENCE of any variant, so a well matching it is a failed insertion, not a
    construct. Named by convention in the supplied maps (e.g. pSL0360_full_uncut)."""
    n = name.lower()
    return "uncut" in n or "_parent" in n or n.endswith("parent")


# The base plasmid a map is built on. Two maps with the same key are the SAME plasmid
# (barcode variants, the uncut parent, an ASO/guide/insert version); different keys are
# different plasmids that merely share the cloning-vector backbone. The identity call is
# only allowed to name a construct from the well's DOMINANT plasmid (the top-breadth
# map's key) — otherwise the short uncut-parent map, which shares every plasmid's
# backbone, wins by the length confound on any well whose real insert wasn't sequenced.
_PLASMID_STEMS = ("psl0360", "psl0423", "htt", "hexa", "exon")


def _plasmid_key(name):
    n = name.lower()
    for stem in _PLASMID_STEMS:
        if stem in n:
            return stem
    return n


def _defining_mask(ri, shortlist, refs):
    """Boolean mask over ref `ri`: 1 at every position covered by a K-mer of `ri` that
    appears in NO other shortlist member. That is `ri`'s DEFINING region — the barcode
    variant, insert, or cargo that, if the reads cover AND match it, proves the well is
    `ri` and not another candidate. Coordinate-free (works across map lengths) and it
    catches insertions/cargo that an alignment-diff would miss. Two identical maps mask
    each other to an empty region — correctly: they are indistinguishable."""
    seq = refs[ri].sequence
    others = set()
    for rj in shortlist:
        if rj == ri:
            continue
        s = refs[rj].sequence
        for p in range(len(s) - K + 1):
            others.add(s[p:p + K])
    mask = bytearray(len(seq))
    for p in range(len(seq) - K + 1):
        if seq[p:p + K] not in others:
            for q in range(p, p + K):
                mask[q] = 1
    return mask


def _score_candidate(ri, refs, dep, bc, mask):
    """Judge one candidate from a pileup of the well's reads onto it. Returns
    (present, defining_id, defining_cov, support):
      defining_cov = fraction of the defining region that got any read;
      defining_id  = fraction of covered defining positions whose consensus base
                     matches this candidate's own sequence;
      support      = total read-bases piled over the defining region (a depth-weighted
                     count, so a thin-but-broadly-covered barcode still clears the floor);
    present = the defining region is COVERED, deeply enough (absolute support), and its
    consensus bases MATCH this candidate. A candidate with no defining region (subset /
    identical to another) is never 'present' — it can only win the nested-map fallback."""
    seq = refs[ri].sequence
    n = len(seq)
    dpos = [p for p in range(n) if mask[p]]
    if len(dpos) < MIN_DEFINING_BP:
        return (False, 0.0, 0.0, 0)
    covered = [p for p in dpos if dep[p] > 0]
    if not covered:
        return (False, 0.0, 0.0, 0)
    defining_cov = len(covered) / len(dpos)
    support = sum(dep[p] for p in covered)
    match = sum(1 for p in covered if bc[p] and max(bc[p], key=bc[p].get) == seq[p])
    defining_id = match / len(covered)
    present = (defining_cov >= DEF_COV_FRAC
               and defining_id >= DEF_ID_MIN
               and support >= DEF_MIN_SUPPORT)
    return (present, defining_id, defining_cov, support)


def _mix_second(win_ri, run_ri, refs, dep_w, bc_w):
    """Is a runner-up REALLY a second construct in the well, not just a sibling inflated
    by shared backbone? Look only at the columns where the runner's sequence differs from
    the winner's, and ask whether the runner's OWN bases actually appear there in the
    reads (piled onto the winner). Shared flanks — which fooled a plain identity compare
    into flagging a false mix — carry no signal here. Needs same-length maps to column-
    align; otherwise conservatively 'no mix'."""
    a = refs[win_ri].sequence
    b = refs[run_ri].sequence
    if len(a) != len(b):
        return False
    diff = [p for p in range(len(a)) if a[p] != b[p] and dep_w[p] > 0]
    if len(diff) < 3:
        return False
    fracs = [bc_w[p].get(b[p], 0) / dep_w[p] for p in diff]
    return sum(fracs) / len(fracs) >= MIX_FRAC


def _parent_identity(pidx, refs, dep, bc):
    """Fraction of covered parent-map positions whose consensus base matches the parent,
    and how many were covered — used to confirm an empty-vector (unmodified parent) call
    once no barcode/insert is present."""
    seq = refs[pidx].sequence
    cov = [p for p in range(len(seq)) if dep[p] > 0 and seq[p] in "ACGT"]
    if not cov:
        return (0.0, 0)
    match = sum(1 for p in cov if bc[p] and max(bc[p], key=bc[p].get) == seq[p])
    return (match / len(cov), len(cov))


def identify_well(reads, refs, pairs, index, banned=frozenset()):
    """Base-level identity, returning an evidence dict (never raises). K-mer breadth
    builds a cheap candidate SHORTLIST; the call itself is made by aligning the reads
    to each shortlisted map and judging its defining region on covered, matching bases
    (see the block above). called=None means no candidate's own sequence is present at
    depth — the insert/barcode was not sequenced, or is not there.

    `banned` = map names this well may not be called as. Used for the second identity
    pass: once the PLATE shows a construct to be background (deep in a few wells, a
    smear everywhere else), the shallow wells are re-identified with it suppressed so
    the well's REAL clone gets scored. Without this the signature scan below returns on
    the contaminant and the true barcode is never even looked at (that is what sent
    C11/260710 grey when its AGGAAGG barcode was sitting there under 122 votes)."""
    readset = _read_kmers(reads)
    cov = _coverage(refs, readset)
    if banned:
        cov = [t for t in cov if refs[t[0]].name not in banned]
    if not cov:
        return dict(called=None, reason="no usable references",
                    candidates=[], best_breadth=0.0, best_idx=None)
    best_b = cov[0][2]
    brmap = {ri: b for ri, c, b in cov}
    disp = lambda idxs: [(refs[ri].name, round(brmap.get(ri, 0.0), 3)) for ri in idxs[:4]]

    # ── UNBIASED PRESENCE SCAN (runs first, before any breadth ranking) ──────────
    # A construct whose OWN map-unique signature is largely present in the reads is
    # really in this well — whatever the sample is labelled, and even if the insert
    # is too long to ever win breadth. This is what catches a clone picked into the
    # WRONG well. Near-identical barcodes have ~no unique signature (they mask each
    # other) so they never trigger here and fall through to the base-level path.
    sig_sets = _signature_sets(refs)
    hits = sorted(((ri, len(s & readset) / len(s)) for ri, s in sig_sets.items()
                   if len(s) >= SIG_MIN_SIZE and len(s & readset) >= SIG_PRESENT_FRAC * len(s)
                   and refs[ri].name not in banned),
                  key=lambda t: -t[1])
    if hits:
        win_ri, win_f = hits[0]
        seq = refs[win_ri].sequence
        sig = sig_sets[win_ri]
        dep, _bc, _, _ = _pileup(pairs[:IDENT_READ_CAP], win_ri, refs, index)
        spos = [p for p in range(len(seq) - K + 1) if seq[p:p + K] in sig]
        supp = sum(dep[p] for p in spos if dep[p] > 0)
        # a second construct on a DIFFERENT base plasmid, also present = a real mix
        others = [(ri, f) for ri, f in hits[1:]
                  if _plasmid_key(refs[ri].name) != _plasmid_key(refs[win_ri].name)]
        run_ri = others[0][0] if others else None
        cand = [(refs[ri].name, round(f, 3)) for ri, f in hits[:4]]
        return dict(called=win_ri, support=supp, runner=run_ri, runner_support=0,
                    mix=bool(others), parent=_is_parent_map(refs[win_ri].name),
                    best_breadth=best_b, candidates=cand, best_idx=win_ri,
                    reason=f"contains {refs[win_ri].name}: {win_f:.0%} of its unique "
                           f"signature is present in the reads")

    if best_b < MIN_IDENTIFY_BREADTH:
        return dict(called=None, best_breadth=best_b, best_idx=cov[0][0],
                    candidates=[(refs[ri].name, round(b, 3)) for ri, c, b in cov[:4]],
                    reason=f"reads reconstruct no supplied map (best {best_b:.0%} covered)")

    # shortlist = top hit's coordinate family ∪ every map within TIER_MARGIN breadth,
    # bounded to SHORTLIST_MAX by breadth (keeps the per-candidate _pileup cost in check)
    fam = _coord_families(refs)
    top_hit = cov[0][0]
    top_fam = fam.get(top_hit)
    shortlist = {ri for ri, c, b in cov if b >= best_b - TIER_MARGIN}
    shortlist |= {ri for ri in range(len(refs))
                  if fam.get(ri) == top_fam and refs[ri].name not in banned}
    shortlist = sorted(shortlist, key=lambda ri: -brmap.get(ri, 0.0))[:SHORTLIST_MAX]

    sample = pairs[:IDENT_READ_CAP]
    scored = []
    piles = {}                              # ri -> (dep, bc), reused by mix + parent tests
    for ri in shortlist:
        dep, bc, _, _ = _pileup(sample, ri, refs, index)
        piles[ri] = (dep, bc)
        mask = _defining_mask(ri, shortlist, refs)
        present, did, dcov, supp = _score_candidate(ri, refs, dep, bc, mask)
        scored.append(dict(ri=ri, present=present, did=did, dcov=dcov, support=supp))

    # best_idx (coverage track) = whoever's defining region got the most coverage; that
    # is the most-relevant map to SHOW even when nothing can be called.
    best_idx = max(scored, key=lambda s: s["dcov"])["ri"] if scored else top_hit
    candidates = disp([s["ri"] for s in sorted(scored, key=lambda s: -s["dcov"])])

    # Only name a construct from the well's DOMINANT plasmid (the top-breadth map's base
    # plasmid). This blocks a well that is primarily some OTHER plasmid — e.g. an HTT-ASO
    # clone whose guide simply wasn't sequenced — from being mislabelled a pSL0360 barcode
    # or the pSL0360 empty vector just because they share the cloning backbone. (Anchoring
    # instead on the deepest defining coverage was tried; it recovers a few empty wells
    # where a superset map wins breadth, but it also turns HTT-with-unsequenced-guide into
    # a confident 'empty' call — the wrong trade when truthfulness beats coverage.)
    top_key = _plasmid_key(refs[top_hit].name)
    dominant = [s for s in scored if _plasmid_key(refs[s["ri"]].name) == top_key]

    # winner = highest base-level identity to its own defining region (that is what tells
    # barcode siblings apart); coverage/support break ties. A background contaminant can't
    # reach here — the DEF_COV_FRAC floor already dropped its half-covered defining region.
    present = sorted((s for s in dominant if s["present"]),
                     key=lambda s: (-s["did"], -s["dcov"], -s["support"]))
    if not present:
        # No barcode/insert present. If the locus WAS sequenced (some sibling's defining
        # region is well covered) yet nothing matched, and the reads rebuild the parent
        # backbone cleanly, the well is the unmodified parent — an empty vector, not a
        # construct. If the locus was NOT sequenced, we honestly can't tell. Only when the
        # DOMINANT plasmid is itself the pSL0360 barcode family (top_key) is 'empty' the
        # right story — an HTT/Exon well missing its own insert is grey, not empty.
        pidx = next((ri for ri in shortlist if _is_parent_map(refs[ri].name)
                     and _plasmid_key(refs[ri].name) == top_key), None)
        sibs = [s for s in dominant if not _is_parent_map(refs[s["ri"]].name)]
        locus = max(sibs, key=lambda s: (s["dcov"], s["support"])) if sibs else None
        if (pidx is not None and locus is not None
                and locus["dcov"] >= DEF_COV_FRAC and locus["support"] >= DEF_MIN_SUPPORT):
            pid, pcov = _parent_identity(pidx, refs, *piles[pidx])
            if pid >= PARENT_ID_MIN and pcov >= PARENT_MIN_COV:
                return dict(called=pidx, support=locus["support"], runner=None,
                            runner_support=0, mix=False, parent=True,
                            best_breadth=best_b, candidates=candidates, best_idx=pidx,
                            reason="unmodified parent — barcode locus sequenced, no "
                                   "insert present")
        return dict(called=None, best_breadth=best_b, candidates=candidates, best_idx=best_idx,
                    reason="reads cover the shared backbone but no construct's own "
                           "sequence is present — the insert/barcode was not sequenced, "
                           "or is not there")
    win = present[0]
    # a runner-up is a real mix only if its OWN distinguishing bases are in the reads
    dep_w, bc_w = piles[win["ri"]]
    mixers = [s for s in present[1:]
              if _mix_second(win["ri"], s["ri"], refs, dep_w, bc_w)]
    run = mixers[0] if mixers else None
    run_ri = run["ri"] if run else None
    run_n = run["support"] if run else 0
    return dict(called=win["ri"], support=win["support"], runner=run_ri, runner_support=run_n,
                mix=bool(run), parent=_is_parent_map(refs[win["ri"]].name),
                best_breadth=best_b, candidates=candidates, best_idx=win["ri"])


def _pileup(pairs, idx, refs, index):
    """Place every read onto reference `idx` (seed-chaining) and return per-position
    depth, base counts, per-base strand counts, and a small-indel tally. `pairs`
    are (seq, qual) tuples; low-quality bases are dropped inside seedchain_tally.
    ic maps (ref_junction, 'ins'|'del', size) -> [fwd, rev, Counter(inserted_seq)]."""
    nn = len(refs[idx].sequence)
    dep = [0] * nn
    bc = [Counter() for _ in range(nn)]
    sc = [dict() for _ in range(nn)]
    ic = {}
    for rd, qual in pairs:
        seedchain_tally(rd, qual, idx, nn, index, dep, bc, sc, ic)
    return dep, bc, sc, ic


def _coverage_track(depth, features, nbins=160):
    """Bin coverage into `nbins` for the HTML coverage bar: each bin is the % of its
    positions that got any read (green = sequenced, red = gap). Also project the
    map's features onto the bar so a gap can be read as 'which part is missing'."""
    n = len(depth)
    if n == 0:
        return [], []
    track = []
    for b in range(nbins):
        i0 = b * n // nbins
        i1 = max(i0 + 1, (b + 1) * n // nbins)
        seg = depth[i0:i1]
        track.append(round(100 * sum(1 for x in seg if x > 0) / len(seg)))
    feats = []
    for f in features:
        if f.kind == "source":
            continue
        s = max(0.0, 100 * f.start / n)
        e = min(100.0, 100 * f.end / n)
        if e - s >= 0.5:
            feats.append([round(s, 1), round(e, 1), f"{f.label} ({f.kind})"])
    feats.sort(key=lambda x: -(x[1] - x[0]))
    return track, feats[:12]


def evaluate_well(res: WellResult, r1: Path, r2: Path, refs, index,
                  distinct_maps, cap: int, banned=frozenset()):
    pairs = list(read_fastq_seqs(r1, cap)) + list(read_fastq_seqs(r2, cap))
    reads = [s for s, _ in pairs]          # sequences only, for k-mer identity
    res.total_reads = len(reads)
    if not reads:
        res.verdict, res.reason = "NO_DATA", "no reads in file"
        return res

    # reads that match ANY supplied map at all — separates "too few reads"
    # (NO_DATA) from "reads don't match anything you supplied" (UNKNOWN).
    mapped = 0
    for rd in reads:
        for o in range(0, len(rd) - K + 1, 5):
            if rd[o:o + K] in index:
                mapped += 1
                break
    res.mapped_reads = mapped
    if mapped < MIN_MAPPED:
        res.verdict = "NO_DATA"
        res.reason = f"only {mapped} reads mapped (<{MIN_MAPPED})"
        return res

    # ── IDENTITY (base-level): k-mer breadth builds a candidate shortlist, then
    # the reads are aligned to each and judged on their DEFINING region — is the
    # construct's own sequence covered at depth (present) AND do the consensus
    # bases match it (identity). Immune to the length confound / mutual masking. ──
    ident = identify_well(reads, refs, pairs, index, banned=banned)
    res.candidates = ident.get("candidates", [])
    res.best_breadth = ident.get("best_breadth", 0.0)
    if ident["called"] is None:
        res.identified = False
        res.called = None
        res.breadth = res.best_breadth
        res.reason = ident.get("reason", "no confident match")
        # still draw a coverage track against the best-matching map so the reader
        # can SEE what was sequenced and what's missing (e.g. an uncovered barcode)
        bi = ident.get("best_idx")
        if bi is not None:
            dep, _, sc, _ = _pileup(pairs, bi, refs, index)
            res.cov_track, res.track_features = _coverage_track(dep, refs[bi].features)
            res.track_len = len(refs[bi].sequence)
            res.track_label = refs[bi].name
            res.depth_fwd = [sum(v[0] for v in p.values()) for p in sc]
            res.depth_rev = [sum(v[1] for v in p.values()) for p in sc]
        decide_verdict(res)
        return res

    called_idx = ident["called"]
    called = refs[called_idx]
    res.identified = True
    res.is_parent = bool(ident.get("parent"))
    res.called = called.name
    res.member_support = ident.get("support") or 0
    win_n, run_n = ident.get("support") or 0, ident.get("runner_support") or 0
    res.called_fraction = win_n / (win_n + run_n) if (win_n + run_n) else 1.0
    if ident.get("mix") and ident.get("runner") is not None:
        res.mix_flag = True
        res.runner_up = refs[ident["runner"]].name
        res.runner_up_fraction = run_n / (win_n + run_n) if (win_n + run_n) else 0.0
        res.reason = (f"two constructs — {res.called} ({win_n} unique reads) "
                      f"+ {res.runner_up} ({run_n})")

    # ── pile reads onto the called construct ──
    # Pile EVERY read; seedchain_tally only places a read that has >=3 seeds
    # matching this construct, so shared backbone covers fully and truly foreign
    # reads are dropped automatically.
    depth, base_counts, strand_counts, indel_counts = _pileup(
        pairs, called_idx, refs, index)
    disc = distinct_maps.get(called_idx, {})
    n = len(called.sequence)
    covered = sum(1 for d in depth if d > 0)
    res.breadth = covered / n if n else 0.0
    nz = [d for d in depth if d > 0]
    res.mean_depth = sum(nz) / len(nz) if nz else 0.0
    res.assayed = bytes(1 if d >= DIFF_MIN_DEPTH else 0 for d in depth)
    # coverage track for the click panel: what was sequenced (green) vs missing (red)
    res.cov_track, res.track_features = _coverage_track(depth, called.features)
    res.track_len = n
    res.track_label = called.name
    # per-position strand depth for the radial view (sum over bases at each position)
    res.depth_fwd = [sum(v[0] for v in sc.values()) for sc in strand_counts]
    res.depth_rev = [sum(v[1] for v in sc.values()) for sc in strand_counts]

    # ── QUALITY: per position, is each nucleotide true? ──
    # Stack the reads; at each position read the consensus. A non-reference base
    # counts only if it's a REAL variant (more reads than error explains). Then:
    #   - reference base no longer real, alt is        -> homozygous MUTATION
    #   - BOTH the reference base and the alt are real  -> HETEROGENEOUS (not pure)
    #   - only the reference is real                    -> clean (skip)
    diffs: list[Diff] = []
    for pos in range(n):
        d = depth[pos]
        if d < DIFF_MIN_DEPTH:
            continue
        exp = called.sequence[pos]
        if exp not in "ACGT":
            continue
        cc = base_counts[pos]
        ref_count = cc[exp]
        alt = max("ACGT", key=lambda b: -1 if b == exp else cc[b])
        alt_count = cc[alt]
        if not _is_real_variant(alt_count, d):
            continue                         # clean, or below the error floor
        f = alt_count / d
        if f >= MUT_FRACTION:
            kind = "mutation"                # alt dominant -> homozygous change
        elif f >= HET_FRACTION:
            kind = "heterogeneous"           # two substantial alleles -> not pure
        else:
            kind = "minor"                   # real, but a low-level subpopulation
        note = ""
        if alt in disc.get(pos, {}).values():
            note = "matches " + next(refs[m].name
                                     for m, mb in disc[pos].items() if mb == alt)
        # strand-bias / oxidation filter: a REAL variant shows on both DNA strands;
        # 8-oxoG damage (G>T on one strand, C>A on the other) and most prep noise are
        # strand-skewed. If the position is covered both ways but the alt is basically
        # one-strand, demote a would-be mutation/het to a non-disqualifying "artifact".
        fwd_c, rev_c = strand_counts[pos].get(alt, (0, 0))
        pf = sum(v[0] for v in strand_counts[pos].values())
        pr = sum(v[1] for v in strand_counts[pos].values())
        if (kind in ("mutation", "heterogeneous") and pf > 0 and pr > 0
                and alt_count >= STRAND_MIN_ALT
                and min(fwd_c, rev_c) / alt_count < STRAND_MIN_FRAC):
            oxid = (exp == "G" and alt == "T") or (exp == "C" and alt == "A")
            kind = "artifact"
            note = (f"strand-biased {exp}>{alt} ({fwd_c}+/{rev_c}-)"
                    + (" — likely 8-oxoG oxidation" if oxid else " — likely prep artifact"))
        # local sequence context (map vs reads) so the change can be read by hand
        cref, cobs, ci = _ctx_window(called.sequence, base_counts, depth, pos, alt)
        # protein-level impact, but only for a change inside a coding-ish feature
        impact = None
        for ft in called.features:
            if ft.start <= pos < ft.end and _CODING_RE.search(ft.label or ""):
                impact = _coding_impact(called.sequence, pos, alt, ft.start, ft.end)
                if impact:
                    break
        diffs.append(Diff(
            pos=pos, expected=exp, observed=alt, depth=d, fraction=alt_count / d,
            kind=kind, features=_features_at(called, pos, pos + 1), note=note,
            ref_count=ref_count, alt_count=alt_count,
            alt_fwd=fwd_c, alt_rev=rev_c,
            ctx_ref=cref, ctx_obs=cobs, ctx_i=ci, ctx_start=pos - ci,
            impact=impact))

    # ── small indels (frameshift-class) ── call them with the SAME error-floor +
    # both-strand bar as substitutions, so Tn5 mapping wobble can't fabricate one.
    # A frameshift in a coding feature is the one defect a substitution pileup
    # misses entirely, so surface it as its own kind of change.
    for (rj, ikind, isize), (ifwd, irev, iseq_cnt) in indel_counts.items():
        support = ifwd + irev
        # a deletion drops depth AT the gap, so read the FLANKING coverage
        jd = max(depth[(rj - 1) % n], depth[rj % n], depth[(rj + isize) % n])
        if jd < DIFF_MIN_DEPTH or not _is_real_variant(support, jd):
            continue                          # below the error floor for this depth
        frac = support / jd if jd else 0.0
        onesided = (support >= STRAND_MIN_ALT
                    and min(ifwd, irev) / support < STRAND_MIN_FRAC)
        # a frameshift only means anything inside an ORF/exon/CDS
        coding = None
        for ft in called.features:
            if ft.start <= rj < ft.end and _CODING_RE.search(ft.label or ""):
                coding = ft
                break
        iseq = iseq_cnt.most_common(1)[0][0] if iseq_cnt else ""
        del_seq = called.sequence[rj:rj + isize]
        signed = isize if ikind == "ins" else -isize
        observed = ("+" + iseq) if ikind == "ins" else ("-" + del_seq)
        expected = "·" if ikind == "ins" else (del_seq or "·")
        if onesided:                          # one-strand indel = prep/mapping artifact
            kind, impact = "artifact", None
            note = (f"strand-biased {isize} bp {ikind} ({ifwd}+/{irev}-) "
                    f"— likely prep/mapping artifact")
        elif coding is not None and isize % 3 != 0:
            kind = "mutation"
            note = f"frameshift ({signed:+d} bp)"
            impact = ("frameshift",
                      f"frameshift {signed:+d} bp — downstream protein scrambled")
        elif coding is not None:
            kind = "mutation"
            note = "in-frame indel"
            impact = ("inframe_indel",
                      f"{signed:+d} bp in-frame — adds/removes residue(s)")
        else:                                 # backbone indel: still a change, but
            kind, impact = "mutation", None   # _func_call flags it backbone-only
            note = f"{isize} bp {ikind}"
        cref, cobs, ci = _ctx_window(called.sequence, base_counts, depth, rj,
                                     called.sequence[rj] if rj < n else "N")
        diffs.append(Diff(
            pos=rj, expected=expected, observed=observed, depth=jd, fraction=frac,
            kind=kind, features=_features_at(called, rj, rj + max(1, isize)),
            note=note, ref_count=max(0, jd - support), alt_count=support,
            alt_fwd=ifwd, alt_rev=irev,
            ctx_ref=cref, ctx_obs=cobs, ctx_i=ci, ctx_start=rj - ci,
            impact=impact, indel_kind=ikind, indel_size=isize, indel_seq=iseq))
    diffs.sort(key=lambda x: x.pos)
    res.differences = diffs

    # ── suspected internal deletion: a dip well below flanking coverage ──
    res.deletion = _deletion_dip(depth)
    if res.deletion:
        s, e, bp = res.deletion
        res.deletion_features = _features_at(called, s, e)

    # runs with NO reads: the corrected .gb inherits these from the map unverified
    res.called_idx = called_idx
    res.uncovered = _uncovered_runs(depth, minlen=20)

    _recount(res)
    decide_verdict(res)
    return res


def _is_real_variant(count, depth):
    """Is `count` reads of one base more than sequencing ERROR explains at this
    depth? Binomial test (normal approximation with continuity correction)
    against ERROR_RATE. This is the truth-test: it trusts 500 reads at 5% AND
    distrusts 4 reads at 13% — count and fraction together, no flat threshold."""
    if count < MIN_ALT_READS:
        return False
    mu = depth * ERROR_RATE
    sd = math.sqrt(depth * ERROR_RATE * (1.0 - ERROR_RATE))
    if sd <= 0:
        return count >= MIN_ALT_READS
    z = (count - 0.5 - mu) / sd
    return z >= Z_SIGNIF


def _recount(res: WellResult):
    # clone-specific homozygous mutations (the disqualifying kind): substitution
    # mutations carry no note; indel mutations DO (e.g. "frameshift") but still
    # condemn the clone, so count them by their indel_kind regardless of note.
    res.n_mut = sum(1 for x in res.differences if x.kind == "mutation"
                    and (x.indel_kind or not x.note))
    # heterogeneous positions: two substantial alleles = not a pure clone
    res.n_het = sum(1 for x in res.differences if x.kind == "heterogeneous")
    res.n_minor = sum(1 for x in res.differences if x.kind == "minor")
    res.n_artifact = sum(1 for x in res.differences if x.kind == "artifact")
    res.n_systematic = sum(1 for x in res.differences if x.kind == "systematic")


def decide_verdict(res: WellResult):
    """Turn the measurements into a PICK decision. Order = priority.
    Key principle: only WELL-SPECIFIC fixed differences (n_fixed) condemn a
    clone. Differences shared across the whole plate (n_systematic) are a map
    error, not a mutation, and never make a well red."""
    # genuinely not identifiable -> say UNKNOWN, never fake a MUT call
    if res.mapped_reads < MIN_MAPPED:
        res.verdict, res.reason = "NO_DATA", res.reason or "too few reads"
        return
    # Identity is set upstream by the two-stage engine: a well is UNKNOWN only
    # when the reads reconstruct no map, or cover the shared backbone but never
    # span the region that names the construct. Low breadth on an IDENTIFIED
    # well is kept and flagged YELLOW below, not thrown out.
    if not res.identified:
        res.verdict, res.called = "RED_UNKNOWN", None
        res.reason = res.reason or "no confident match to any supplied map"
        return
    if res.is_parent:
        res.verdict = "RED_EMPTY"
        res.reason = ("unmodified parent / empty vector — backbone present but no "
                      "insert or barcode (failed clone)")
        return
    if res.mix_flag:
        res.verdict = "RED_MIX"
        return
    if res.deletion and res.deletion[2] >= 150:
        s, e, bp = res.deletion
        feat = ", ".join(res.deletion_features[:3]) or f"bp {s+1}-{e}"
        res.verdict, res.reason = "RED_BROKEN", f"{bp} bp deletion ({feat})"
        return
    if res.n_mut > 0:
        muts = [x for x in res.differences if x.kind == "mutation"
                and (x.indel_kind or not x.note)]
        # lead with the worst: a frameshift in a critical feature is the headline
        ex = max(muts, key=lambda x: (bool(x.indel_kind),
                                      _feat_cat(x.features) == "critical",
                                      x.fraction))
        res.verdict = "RED_MUT"
        if ex.indel_kind:
            feat = (ex.features[0].split(" (")[0] if ex.features else "")
            what = (ex.note or f"{ex.indel_size} bp {ex.indel_kind}")
            res.reason = (f"{res.n_mut} clone-specific mutation(s), "
                          f"e.g. {what} at pos {ex.pos+1}"
                          + (f" in {feat}" if feat else "")
                          + f" ({ex.fraction:.0%} of {ex.depth}x)")
        else:
            res.reason = (f"{res.n_mut} clone-specific mutation(s), "
                          f"e.g. pos {ex.pos+1} {ex.expected}>{ex.observed} "
                          f"({ex.fraction:.0%} of {ex.depth}x)")
        return
    if res.n_het > 0:
        ex = next(x for x in res.differences if x.kind == "heterogeneous")
        res.verdict = "RED_HET"
        res.reason = (f"{res.n_het} heterogeneous position(s) — two alleles, not a "
                      f"pure clone, e.g. pos {ex.pos+1} "
                      f"{ex.expected}={ex.ref_count}/{ex.observed}={ex.alt_count}")
        return
    if res.breadth < BREADTH_GREEN:
        res.verdict = "YELLOW"
        res.reason = f"partial coverage (breadth {res.breadth:.0%})"
        return
    if res.mean_depth < THIN_DEPTH:
        res.verdict = "YELLOW"
        res.reason = f"thin coverage (mean {res.mean_depth:.0f}x)"
        return
    res.verdict = "GREEN"
    notes = []
    if res.n_systematic:
        notes.append(f"{res.n_systematic} systematic map mismatch(es)")
    if res.n_minor:
        notes.append(f"{res.n_minor} low-level variant(s) <30%")
    res.reason = "matches the map" + (f"  ({'; '.join(notes)} ignored)"
                                      if notes else "")


def _coord_families(refs):
    """ref_idx -> family_id for maps that SHARE A COORDINATE FRAME: identical length and
    near-identical sequence (barcode siblings off one parent). Maps of different lengths
    never pool — position 4533 in one is not position 4533 in the other.

    This is not the same question as `group_refs` asks. That one chains anything sharing
    a vector backbone, which is most of a match/ folder; pooling by it would compare
    positions across maps whose coordinates do not correspond."""
    fam, fid = {}, 0
    by_len = defaultdict(list)
    for ri, r in enumerate(refs):
        by_len[len(r.sequence)].append(ri)
    for _, group in sorted(by_len.items()):
        pending = list(group)
        while pending:
            seed = pending.pop(0)
            fam[seed] = fid
            s = refs[seed].sequence
            rest = []
            for ri in pending:
                t = refs[ri].sequence
                if sum(1 for a, b in zip(s, t) if a != b) <= COORD_FAMILY_MAXDIFF * len(s):
                    fam[ri] = fid
                else:
                    rest.append(ri)
            pending = rest
            fid += 1
    return fam


def find_background_maps(results):
    """Plate-wide pass: which constructs are BACKGROUND rather than this well's clone?

    A construct CALLED in many wells but deep in only a few is a plate-wide contaminant
    (pSL0423 on these plates — a high-copy plasmid smearing its unique guide everywhere).
    Returns {map_name: (max_depth, floor)}: below `floor` a call of that map is the smear,
    not the clone. A real, localised construct (an Exon swapped into 1-3 wells) is called
    in too few wells to trip this, so it is never flagged.

    NOTE this only IDENTIFIES background. It deliberately does not decide any well's fate:
    a well whose top call is background still has a real clone in it most of the time, and
    that clone deserves to be scored (see the second identity pass in run()). The earlier
    version of this pass demoted such wells straight to 'can't tell', which threw away good
    clones — C11/260710 was really an AGGAAGG barcode sitting under a 33x pSL0423 smear."""
    by_call = defaultdict(list)
    for r in results:
        if r.called and not _is_parent_map(r.called):
            by_call[r.called].append(r)
    bg = {}
    for construct, wells in by_call.items():
        if len(wells) < CONTAM_MIN_WELLS:
            continue
        maxd = max(r.mean_depth for r in wells)
        floor = max(CONTAM_DEPTH_FRAC * maxd, CONTAM_ABS_DEPTH)
        if any(r.mean_depth < floor for r in wells):
            bg[construct] = (maxd, floor)
    return bg


def shallow_background_wells(results, bg):
    """Indices of wells whose call is only the background smear — the ones worth
    re-identifying with that map suppressed."""
    out = []
    for i, r in enumerate(results):
        if r.called in bg and r.mean_depth < bg[r.called][1]:
            out.append(i)
    return out


def mark_systematic(results, refs):
    """Second pass across the whole plate: a fixed difference that shows up in most wells
    of a construct FAMILY is the MAP being wrong (the .gb export is stale), not the clones
    being mutated — siblings share the same backbone, so they share the same map errors.

    Pool by family, not by individual sibling: a construct picked into one or two wells can
    never reach a stable threshold on its own, and the errors live in shared backbone
    anyway, so they recur at the SAME position across every sibling.

    Judge each position only against the wells that could actually SEE it. Coverage is
    patchy, so counting wells that never reached a position as evidence against a map error
    there silently vetoes real ones."""
    by_name = {r.name: i for i, r in enumerate(refs)}
    fam = _coord_families(refs)
    by_fam = defaultdict(list)
    for r in results:
        if r.called and not r.mix_flag:
            ri = by_name.get(r.called)
            if ri is not None:
                by_fam[fam[ri]].append(r)
    for _, wells in by_fam.items():
        pos_count = Counter()
        for r in wells:
            for d in r.differences:
                # any real call recurring across the family's wells is a map /
                # sequence-context artifact, not a per-clone event (indels carry a
                # note but vote too — a plate-wide indel is a stale map, not a clone).
                # A position already noted as matching a sibling is excluded: that is
                # cross-contamination, and the barcode is where siblings must differ.
                if d.kind in ("mutation", "heterogeneous", "minor",
                              "systematic") and (d.indel_kind or not d.note):
                    pos_count[(d.pos, d.observed)] += 1
        systematic = set()
        for (pos, obs), c in pos_count.items():
            seen = sum(1 for r in wells if pos < len(r.assayed) and r.assayed[pos])
            if seen < 2:
                continue                      # too few wells judged it to conclude anything
            if c >= max(2, int(round(SYSTEMATIC_FRACTION * seen))):
                systematic.add((pos, obs))
        for r in wells:
            for d in r.differences:
                if ((d.pos, d.observed) in systematic
                        and (d.indel_kind or not d.note)):
                    d.kind = "systematic"
            _absorb_adjacent_systematic(r)
            _recount(r)
            decide_verdict(r)


def _absorb_adjacent_systematic(res: WellResult):
    """A local map/reference discrepancy often spans a short RUN of adjacent bases
    (the .gb export is wrong for 2-3 nt in a row), which surfaces as adjacent
    fully-fixed substitutions. If the plate-wide vote confirmed one of them as
    systematic but a neighbour was only seen in this well (a sibling lacked depth
    at that exact base), the neighbour is the SAME map error — not a clone-specific
    mutation. Absorb any near-fixed change immediately adjacent (<=1 nt) to a
    systematic one into 'systematic', propagating along the run."""
    fixed = [d for d in res.differences
             if not d.note and d.fraction >= 0.95 and d.kind in
             ("mutation", "heterogeneous", "systematic")]
    fixed.sort(key=lambda d: d.pos)
    changed = True
    while changed:
        changed = False
        for i, d in enumerate(fixed):
            if d.kind != "systematic":
                continue
            for j in (i - 1, i + 1):
                if 0 <= j < len(fixed):
                    nb = fixed[j]
                    if nb.kind != "systematic" and abs(nb.pos - d.pos) <= 1:
                        nb.kind = "systematic"
                        changed = True


def _features_at(ref: Reference, start: int, end: int):
    out = []
    for f in ref.features:
        if f.start < end and f.end > start:
            lab = f"{f.label} ({f.kind})"
            if lab not in out:
                out.append(lab)
    return out


def _deletion_dip(depth):
    """Find the longest internal run of NEAR-ZERO coverage — a real dropout, not
    a Tn5 coverage valley. Requires the run to be essentially empty (<5% of
    median AND <=2 reads) and flanked by well-covered sequence. Returns
    (start,end,bp) or None."""
    n = len(depth)
    nz = [d for d in depth if d > 0]
    if not nz:
        return None
    med = sorted(nz)[len(nz) // 2]
    if med < 20:                       # too shallow overall to judge a dropout
        return None
    thresh = min(2, 0.05 * med)        # near-zero, not merely "below average"
    best = None
    i = 0
    while i < n:
        if depth[i] <= thresh:
            j = i
            while j < n and depth[j] <= thresh:
                j += 1
            flank_ok = (i > 50 and j < n - 50
                        and depth[i - 1] >= 0.3 * med and depth[j] >= 0.3 * med)
            if flank_ok and (best is None or (j - i) > best[2]):
                best = (i, j, j - i)
            i = j
        else:
            i += 1
    return best if best and best[2] >= DELETION_MIN_BP else None


def _uncovered_runs(depth, minlen=20):
    """Runs of >= minlen consecutive positions with NO reads. The corrected .gb
    keeps the map bases over these (we didn't sequence them) but marks them so the
    reader knows they're inherited, not confirmed. Returns [(start, end)]."""
    runs, n, i = [], len(depth), 0
    while i < n:
        if depth[i] == 0:
            j = i
            while j < n and depth[j] == 0:
                j += 1
            if j - i >= minlen:
                runs.append((i, j))
            i = j
        else:
            i += 1
    return runs


# ─────────────────────────────────────────────────────────────────────────────
# 8b.  Corrected GenBank — the read-consensus of the called map, downloadable
# ─────────────────────────────────────────────────────────────────────────────

def _gb_origin(seq):
    """Sequence block in GenBank ORIGIN format: 60 bp/line, 10-bp groups, the
    1-based position of the first base of each line in a right-justified column."""
    out, s = [], seq.lower()
    for i in range(0, len(s), 60):
        chunk = s[i:i + 60]
        groups = " ".join(chunk[j:j + 10] for j in range(0, len(chunk), 10))
        out.append(f"{i + 1:>9} {groups}")
    return "\n".join(out)


def _gb_feature(kind, start, end, label):
    """One FEATURES entry (1-based inclusive location) with a /label qualifier."""
    loc = f"{start + 1}..{end}"
    lab = str(label).replace('"', "'")
    return (f"     {kind:<15} {loc}\n"
            f"                     /label=\"{lab}\"")


def build_corrected_gb(res: WellResult, ref: Reference) -> str:
    """The corrected transcript: start from the CALLED map, then apply every
    confident change the reads actually showed — fixed substitutions, real small
    indels, and stale-map ('systematic') edits — and take the majority base at a
    heterogeneous site. Runs with no coverage are copied from the map and flagged
    'unverified_low_coverage', because we did not sequence them. This is a read
    consensus on the map, NOT a de-novo assembly."""
    seq = list(ref.sequence)
    feats = [Feature(f.kind, f.label, f.start, f.end) for f in ref.features]
    qc = []  # (kind, start, end, label) QC annotations added below

    # collect the confident edits (0-based). Apply substitutions/HET in place;
    # collect indels to apply 3'->5' so earlier coordinates don't shift.
    indels = []                          # (pos, ikind, size, seq)
    for d in res.differences:
        if d.indel_kind and (d.kind == "mutation" or d.kind == "systematic"):
            indels.append((d.pos, d.indel_kind, d.indel_size, d.indel_seq))
            qc.append(("misc_feature", d.pos, d.pos + max(1, d.indel_size),
                       f"corrected_indel {d.observed}"
                       f"{' ' + d.note if d.note else ''}"))
        elif d.kind in ("mutation", "systematic") and d.fraction >= 0.5:
            if 0 <= d.pos < len(seq) and d.observed and d.observed[0] in "ACGT":
                seq[d.pos] = d.observed[0]
                qc.append(("misc_feature", d.pos, d.pos + 1,
                           f"corrected_substitution {d.expected}>{d.observed}"))
        elif d.kind == "heterogeneous":
            # not a pure clone — record the majority allele so the sequence is usable
            maj = d.observed if d.alt_count >= d.ref_count else d.expected
            if 0 <= d.pos < len(seq) and maj and maj[0] in "ACGT":
                seq[d.pos] = maj[0]
                qc.append(("misc_feature", d.pos, d.pos + 1,
                           f"heterozygous_majority {d.expected}/{d.observed}"))

    for s, e in res.uncovered:
        qc.append(("misc_feature", s, e, "unverified_low_coverage"))

    # apply indels 3'->5'; shift feature/QC coordinates that sit AFTER each indel
    all_feats = feats + [Feature(k, lab, s, e) for (k, s, e, lab) in qc]
    for pos, ikind, size, iseq in sorted(indels, key=lambda x: -x[0]):
        if ikind == "ins":
            seq[pos:pos] = list(iseq)
            delta = size
        else:
            del seq[pos:pos + size]
            delta = -size
        for f in all_feats:
            if f.start >= pos:
                f.start += delta
            if f.end > pos:
                f.end += delta
    corrected = "".join(seq)

    name = ("corrected_" + (res.called or ref.name))[:60].replace(" ", "_")
    header = [
        f"LOCUS       {name:<24} {len(corrected)} bp    DNA     linear   "
        f"     SYN",
        f"DEFINITION  pickme read-consensus of {res.well} on {res.called or ref.name}.",
        "COMMENT     Generated by pickme. Consensus of this well's reads on the "
        "called map;",
        "            'unverified_low_coverage' spans had no reads and are copied "
        "from the map.",
        "FEATURES             Location/Qualifiers",
    ]
    body = [_gb_feature(f.kind or "misc_feature", f.start, f.end, f.label)
            for f in all_feats if 0 <= f.start < f.end <= len(corrected)]
    return ("\n".join(header + body)
            + "\nORIGIN\n" + _gb_origin(corrected) + "\n//\n")


# ─────────────────────────────────────────────────────────────────────────────
# 9.  Discovery
# ─────────────────────────────────────────────────────────────────────────────

WELL_RE = re.compile(r"([A-H])(\d{1,2})")
PLATE_RE = re.compile(r"(aRY\d+|plate[_-]?\w+)", re.IGNORECASE)


def discover(root: Path, fastq_dir, refs_dir):
    rs = refs_dir or root
    ref_files = [p for p in rs.rglob("*") if p.suffix.lower() in (".gb", ".dna")]
    fs = fastq_dir or root
    r1_files = sorted(p for p in fs.rglob("*")
                      if p.name.endswith((".fastq.gz", ".fastq")) and "_R1" in p.name)
    return ref_files, r1_files


def parse_sample(r1: Path):
    stem = r1.name
    for ext in (".fastq.gz", ".fastq"):
        if stem.endswith(ext):
            stem = stem[:-len(ext)]
    stem = re.sub(r"_R1.*$", "", stem)
    wm = WELL_RE.findall(stem)
    well = (wm[-1][0] + wm[-1][1]) if wm else "?"
    pm = PLATE_RE.search(stem)
    plate = pm.group(1) if pm else "plate1"
    return stem, plate, well


def find_r2(r1: Path):
    c = r1.parent / r1.name.replace("_R1", "_R2")
    return c if c.exists() else None


def intended_label(sample: str) -> str:
    """The construct name the way it appears on the bench plate map (what was
    picked into this well), cleaned from the FASTQ sample name so the grid can be
    copy-pasted straight into the lab's Excel sheet."""
    base = re.split(r"_aRY|_S\d+_|_S\d+$", sample)[0]
    base = re.sub(r"_SS_(\d)_(\d)(_|$)", r"_SS+\1.\2\3", base)   # _SS_4_5_ -> _SS+4.5_
    base = base.replace("Exon4_with_EF1A", "Exon4 with EF1A")
    return base


# ─────────────────────────────────────────────────────────────────────────────
# 10.  Rendering
# ─────────────────────────────────────────────────────────────────────────────

_ACTION = {
    "GREEN": "Use — matches the map.",
    "YELLOW": "Check — partial/thin coverage.",
    "RED_MUT": "Clone has a real mutation — discard or re-pick.",
    "RED_HET": "Not a pure clone (two alleles) — re-streak and re-pick.",
    "RED_MIX": "Two constructs present — re-streak and re-sequence.",
    "RED_BROKEN": "Large deletion — discard.",
    "RED_EMPTY": "Empty vector (unmodified parent) — insertion failed, re-pick.",
    "RED_UNKNOWN": "Not identified — check the reference folder / contamination.",
    "NO_DATA": "Re-sequence — not enough reads.",
}
_LABEL = {
    "GREEN": "MATCH", "YELLOW": "CHECK", "RED_MUT": "RED — MUTATION",
    "RED_HET": "RED — HETEROGENEOUS", "RED_MIX": "RED — MIXED WELL",
    "RED_BROKEN": "RED — DELETION", "RED_EMPTY": "RED — EMPTY VECTOR",
    "RED_UNKNOWN": "RED — UNKNOWN",
    "NO_DATA": "NO DATA",
}


def text_block(r: WellResult) -> str:
    L = [f"{r.plate}  {r.well}  —  {r.called or 'UNKNOWN'}  ·  {_LABEL[r.verdict]}",
         f"  sample: {r.sample}",
         f"  reads: {r.total_reads:,} total, {r.mapped_reads:,} mapped "
         f"({(r.mapped_reads/r.total_reads*100 if r.total_reads else 0):.0f}%)"]
    if r.called:
        L.append(f"  construct: {r.called} ({r.called_fraction:.0%} of mapped reads)")
    if r.runner_up:
        L.append(f"  also present: {r.runner_up} ({r.runner_up_fraction:.0%})")
    for name, d, md in r.background:
        L.append(f"  background: {name} is in this well at {d:.0f}x, but runs {md:.0f}x "
                 f"where it is the real clone — read as contamination, not this pick")
    L.append(f"  coverage: {r.breadth:.0%} of plasmid, mean {r.mean_depth:.0f}x")
    mut = [d for d in r.differences if d.kind == "mutation"
           and (d.indel_kind or not d.note)]
    het = [d for d in r.differences if d.kind == "heterogeneous"]
    systematic = [d for d in r.differences if d.kind == "systematic"]
    sibm = [d for d in r.differences
            if d.note and not d.indel_kind and d.kind != "heterogeneous"]
    if mut:
        L.append(f"  MUTATIONS ({len(mut)}) — consensus differs from the map "
                 f"(homozygous, above the error floor):")
        for d in mut[:30]:
            ft = f"  [{', '.join(d.features)}]" if d.features else ""
            if d.indel_kind:                 # small indel: show it as an indel
                what = (f"{d.observed} ({d.indel_size} bp "
                        f"{'insertion' if d.indel_kind == 'ins' else 'deletion'}"
                        f"{', ' + d.note if d.note else ''})")
                L.append(f"     pos {d.pos+1:<6} {what}  "
                         f"{d.alt_count}/{d.depth} reads ({d.fraction:.0%}){ft}")
            else:
                L.append(f"     pos {d.pos+1:<6} {d.expected}>{d.observed}  "
                         f"{d.alt_count}/{d.depth} reads ({d.fraction:.0%}){ft}")
        if len(mut) > 30:
            L.append(f"     ... and {len(mut)-30} more")
    if het:
        L.append(f"  HETEROGENEOUS positions ({len(het)}) — two real alleles, "
                 f"this colony isn't a single clone:")
        for d in het[:30]:
            ft = f"  [{', '.join(d.features)}]" if d.features else ""
            nb = f"  ({d.note})" if d.note else ""
            L.append(f"     pos {d.pos+1:<6} {d.expected}={d.ref_count} / "
                     f"{d.observed}={d.alt_count}  of {d.depth}x{ft}{nb}")
    minor = [d for d in r.differences if d.kind == "minor"]
    if minor:
        L.append(f"  low-level variants ({len(minor)}, <30% of reads) — real but "
                 f"a minor subpopulation; not disqualifying")
    if systematic:
        L.append(f"  systematic map mismatches ({len(systematic)}) — present in "
                 f"most wells, so the .gb map is off here, NOT this clone "
                 f"(ignored for the verdict)")
    if sibm:
        L.append(f"  positions matching a sibling construct ({len(sibm)}) — "
                 f"possible minor cross-contamination")
    if r.deletion:
        s, e, bp = r.deletion
        ft = f"  [{', '.join(r.deletion_features)}]" if r.deletion_features else ""
        L.append(f"  suspected deletion: ~{bp} bp of low coverage at "
                 f"{s+1}-{e}{ft}")
    L.append(f"  VERDICT: {r.verdict}" + (f" — {r.reason}" if r.reason else ""))
    L.append(f"  action: {_ACTION[r.verdict]}")
    return "\n".join(L)


def write_text(results, path):
    Path(path).write_text(("\n" + "─" * 64 + "\n").join(text_block(r) for r in results),
                          encoding="utf-8")


def write_json(results, path):
    out = []
    for r in results:
        d = asdict(r)
        d.pop("assayed", None)      # per-position depth mask: working state, not a result
        out.append(d)
    # Compact, not indented: the per-position depth_fwd/depth_rev tracks are thousands of
    # ints per well, and indent=2 would put each on its own line (a multi-MB, million-line
    # file). This is a machine-read artifact — the .txt/.html are the human-readable ones.
    Path(path).write_text(json.dumps(out, separators=(",", ":")), encoding="utf-8")


# ── HTML report ──────────────────────────────────────────────────────────────
# Verdict = raw match vs the map.  Functional call = does the difference land in
# a region that changes the construct, or only in vector backbone?  A "MUT" whose
# only edits are in the ori / marker is still pickable — that distinction is the
# whole point of this page.

_HTML_VMAP = {   # verdict -> (short tag, css class, plain meaning)
    "GREEN":       ("PICK",  "v-green",  "Clean &amp; correct"),
    "YELLOW":      ("CHECK", "v-yellow", "Minor concern"),
    "RED_MUT":     ("MUT",   "v-mut",    "Right construct, has mutation(s)"),
    "RED_HET":     ("HET",   "v-het",    "Mixed colony (two alleles)"),
    "RED_BROKEN":  ("DEL",   "v-broken", "Internal deletion"),
    "RED_MIX":     ("MIX",   "v-mix",    "Two constructs in one well"),
    "RED_EMPTY":   ("EMPTY", "v-empty",  "Empty vector — insertion failed"),
    "RED_UNKNOWN": ("UNK",   "v-unk",    "Matches nothing supplied"),
    "NO_DATA":     ("&mdash;", "v-none", "No usable reads"),
}
_HTML_ORDER = ["GREEN", "YELLOW", "RED_MUT", "RED_HET", "RED_BROKEN",
               "RED_MIX", "RED_EMPTY", "RED_UNKNOWN", "NO_DATA"]

# Features that ARE the construct (vs vector plumbing).  Tune to taste — this is
# the one editorial call on the page, and it drives the pickable/re-pick split.
_CRIT_RE = re.compile(
    r"exon|intron|splice|payload|cargo|hhr|ribozyme|ef1a|eblock|ebf|hyb|"
    r"acceptor|donor|branch|mecp2|\bss\b|\bog\b|insert|transgene|kozak|"
    r"\bcds\b|orf|linker|guide|target", re.I)


def _short(name):
    return (name or "").replace("pSL0360_", "").replace("_FULL", "")


def _feat_cat(feats):
    if not feats:
        return "other"
    return "critical" if any(_CRIT_RE.search(f) for f in feats) else "backbone"


def _func_call(r):
    """(status, badge, detail) — functional verdict, not raw match."""
    v = r.verdict
    if v in ("GREEN", "YELLOW"):
        return ("pick", "PICK", "clean")
    if v == "RED_UNKNOWN" or v == "NO_DATA":
        return ("none", "&mdash;", "no match")
    if v == "RED_MIX":
        return ("repick", "MIX", "two constructs")
    if v == "RED_EMPTY":
        return ("repick", "EMPTY", "empty vector — insertion failed")
    crit, bb = [], 0
    for d in r.differences:
        if d.kind not in ("mutation", "heterogeneous"):
            continue
        if _feat_cat(d.features) == "critical":
            fn = (d.features or ["?"])[0].split(" (")[0]
            if d.indel_kind:
                fs = "frameshift" if "frameshift" in d.note else f"{d.indel_size}bp indel"
                crit.append(f"{d.pos + 1} {d.observed} {fs} in {fn}")
            else:
                crit.append(f"{d.pos + 1} {d.expected}&rarr;{d.observed} in {fn}")
        elif _feat_cat(d.features) == "backbone":
            bb += 1
    if v == "RED_BROKEN":
        return ("repick", "DEL", "internal deletion" + ("; " + crit[0] if crit else ""))
    if crit:
        return ("repick", "HIT", "; ".join(crit[:3]))
    return ("pick", "OK*", f"backbone-only ({bb} change{'s' if bb != 1 else ''})")


def _decision(r):
    """Collapse every well to the only question that matters at the bench:
    can I use it or not? Returns (state, mark, perfect) where state is
    'good' | 'bad' | 'unknown'. The mark is a small annotation (★ perfect,
    a note on a usable-but-imperfect well, or the reason a bad one is bad) —
    it decorates the cell, it never changes the good/bad colour."""
    fstatus = _func_call(r)[0]
    if fstatus == "none":
        return ("unknown", "?", False)
    if fstatus == "repick":                       # not usable — mark says why
        mark = {"RED_MIX": "MIX", "RED_BROKEN": "DEL", "RED_EMPTY": "MT",
                "RED_MUT": "MUT", "RED_HET": "HET"}.get(r.verdict, "NO")
        return ("bad", mark, False)
    # usable
    # ★ = reference-grade: identified, well covered, and NOTHING specific to this
    # clone — no mutation, no heterogeneity, no minor subpopulation, no artifact.
    # The only differences allowed are plate-wide 'systematic' map staleness.
    if (r.verdict == "GREEN" and not r.n_mut and not r.n_het
            and not r.n_minor and not r.n_artifact and r.breadth >= BREADTH_GREEN):
        return ("good", "★", True)                # flawless, like F7
    if r.verdict == "YELLOW":
        return ("good", "~", False)               # usable, partial/thin coverage
    if r.n_het:
        return ("good", "het", False)             # heterogeneous, but only in backbone
    if r.n_mut:
        return ("good", "bb", False)              # a change, but only in backbone
    # clean call, but has low-level background (minors/artifacts) -> quiet dot, no star
    return ("good", "·", False)


_HTML_CSS = r"""
:root{--bg:#f6f7f9;--panel:#fff;--ink:#161a20;--ink2:#5a6472;--line:#e3e7ec;--accent:#0e7c86;
--green:#1f9d57;--green-bg:#e3f5eb;--yellow:#b7791f;--yellow-bg:#fbf1dc;--mut:#d1495b;--mut-bg:#fbe4e7;
--het:#c2410c;--het-bg:#fbe8dc;--broken:#9d174d;--broken-bg:#fadfe9;--mix:#6d28d9;--mix-bg:#ece2fb;
--empty:#8a5a2b;--empty-bg:#f2e9dd;--unk:#8a94a3;--unk-bg:#eef1f4;--none:#b6bdc7;--none-bg:#f2f4f6;}
@media(prefers-color-scheme:dark){:root{--bg:#0f1216;--panel:#171b21;--ink:#e8ecf1;--ink2:#9aa4b2;--line:#262c34;--accent:#3fd0dd;
--green:#3ec77e;--green-bg:#12331f;--yellow:#e0b64d;--yellow-bg:#33280e;--mut:#f2768a;--mut-bg:#3a1620;
--het:#f2955a;--het-bg:#38200f;--broken:#f2679f;--broken-bg:#3a1226;--mix:#a883f2;--mix-bg:#241436;
--empty:#cba06b;--empty-bg:#2c2116;--unk:#8894a3;--unk-bg:#1e242c;--none:#5b6470;--none-bg:#191e24;}}
:root[data-theme="light"]{--bg:#f6f7f9;--panel:#fff;--ink:#161a20;--ink2:#5a6472;--line:#e3e7ec;--accent:#0e7c86;--green:#1f9d57;--green-bg:#e3f5eb;--yellow:#b7791f;--yellow-bg:#fbf1dc;--mut:#d1495b;--mut-bg:#fbe4e7;--het:#c2410c;--het-bg:#fbe8dc;--broken:#9d174d;--broken-bg:#fadfe9;--mix:#6d28d9;--mix-bg:#ece2fb;--empty:#8a5a2b;--empty-bg:#f2e9dd;--unk:#8a94a3;--unk-bg:#eef1f4;--none:#b6bdc7;--none-bg:#f2f4f6;}
:root[data-theme="dark"]{--bg:#0f1216;--panel:#171b21;--ink:#e8ecf1;--ink2:#9aa4b2;--line:#262c34;--accent:#3fd0dd;--green:#3ec77e;--green-bg:#12331f;--yellow:#e0b64d;--yellow-bg:#33280e;--mut:#f2768a;--mut-bg:#3a1620;--het:#f2955a;--het-bg:#38200f;--broken:#f2679f;--broken-bg:#3a1226;--mix:#a883f2;--mix-bg:#241436;--empty:#cba06b;--empty-bg:#2c2116;--unk:#8894a3;--unk-bg:#1e242c;--none:#5b6470;--none-bg:#191e24;}
*{box-sizing:border-box}
body{margin:0;background:var(--bg);color:var(--ink);font-family:ui-sans-serif,-apple-system,"Segoe UI",Roboto,sans-serif;line-height:1.5;-webkit-font-smoothing:antialiased}
.mono{font-family:ui-monospace,"SF Mono",Menlo,monospace;font-variant-numeric:tabular-nums}
article{max-width:1080px;margin:0 auto;padding:40px 24px 80px}
.eyebrow{font-size:12px;letter-spacing:.14em;text-transform:uppercase;color:var(--accent);font-weight:600}
.hd h1{font-size:34px;line-height:1.1;margin:.25em 0 .1em;letter-spacing:-.01em;text-wrap:balance}
.sub{color:var(--ink2);max-width:70ch;margin:.4em 0 1.4em;font-size:15px}
.tiles{display:grid;grid-template-columns:repeat(auto-fit,minmax(112px,1fr));gap:10px;margin:0 0 14px}
.tile{background:var(--panel);border:1px solid var(--line);border-radius:12px;padding:12px 14px;border-left:4px solid var(--unk)}
.tile .tnum{font-size:26px;font-weight:700;font-variant-numeric:tabular-nums;line-height:1}
.tile .tlbl{font-size:11px;letter-spacing:.08em;text-transform:uppercase;font-weight:600;margin-top:4px}
.tile .thuman{font-size:11px;color:var(--ink2);margin-top:2px}
.headline{background:var(--panel);border:1px solid var(--line);border-radius:12px;padding:12px 16px}
.headline .hl-main{font-size:15px;color:var(--ink2)}.headline .hl-main b{color:var(--ink);font-size:19px}
.headline .hl-sub{font-size:12.5px;color:var(--ink2);margin-top:6px;max-width:82ch}.headline .hl-sub b{color:var(--ink)}
.hl-sep{opacity:.4;margin:0 4px}
.legend{display:flex;flex-wrap:wrap;gap:6px 16px;margin:26px 0 18px;font-size:12px;color:var(--ink2)}
.lg{display:inline-flex;align-items:center;gap:6px}
.lg i{width:11px;height:11px;border-radius:3px;display:inline-block;background:var(--unk);border:1px solid rgba(0,0,0,.12)}
.plates{display:flex;flex-direction:column;gap:26px}
.plate h3{margin:0 0 10px;font-size:15px;font-family:ui-monospace,monospace;letter-spacing:.02em}
.pcount{color:var(--ink2);font-weight:400;font-size:12px;margin-left:8px}
.grid{display:grid;grid-auto-flow:row;gap:5px;overflow-x:auto;grid-template-columns:26px repeat(var(--cols,7),minmax(78px,1fr))}
.ghead{font-size:11px;color:var(--ink2);text-align:center;font-family:ui-monospace,monospace;align-self:center}
.ghead.rlab{align-self:stretch;display:flex;align-items:center;justify-content:center}
.cell{border-radius:9px;padding:7px 8px;min-height:52px;border:1px solid var(--line);background:var(--panel);display:flex;flex-direction:column;gap:2px;position:relative;cursor:default}
.cell.empty{background:transparent;border:1px dashed var(--line);opacity:.35}
/* three states only: good (use) · bad (don't use) · unknown */
.cell.d-good{background:var(--green-bg);border-color:var(--green);border-left:4px solid var(--green)}
.cell.d-bad{background:var(--mut-bg);border-color:var(--mut);border-left:4px solid var(--mut)}
.cell.d-unknown{background:var(--unk-bg);border-color:var(--line);border-left:4px solid var(--unk)}
.tile.d-good{border-left-color:var(--green)}.tile.d-bad{border-left-color:var(--mut)}.tile.d-unknown{border-left-color:var(--unk)}
.lg.d-good i{background:var(--green)}.lg.d-bad i{background:var(--mut)}.lg.d-unknown i{background:var(--unk)}
.w-mark{position:absolute;top:4px;right:6px;font-size:10px;font-weight:800;letter-spacing:.02em}
.w-mark.star{font-size:14px;color:#e6a700;top:2px}
.d-good .w-mark{color:var(--green)}.d-bad .w-mark{color:var(--mut)}.d-unknown .w-mark{color:var(--unk2,var(--ink2))}
.cell:hover{outline:2px solid var(--accent);outline-offset:1px;z-index:3}
.w-tag{position:absolute;top:5px;right:7px;font-size:9px;font-weight:700;letter-spacing:.04em;opacity:.85}
.w-well{font-family:ui-monospace,monospace;font-size:11px;color:var(--ink2)}
.w-call{font-size:11px;font-weight:600;line-height:1.15;word-break:break-word}
.w-call.w-dim{color:var(--ink2);font-weight:500}
.fdot{position:absolute;bottom:5px;right:6px;font-size:8.5px;font-weight:800;letter-spacing:.03em;padding:1px 4px;border-radius:5px;opacity:.9}
.f-pick{background:var(--green-bg);color:var(--green)}.f-repick{background:var(--mut-bg);color:var(--mut)}
.cell.fn-repick{box-shadow:inset 0 0 0 1.5px var(--mut)}
.v-green{border-left-color:var(--green)}.cell.v-green{background:var(--green-bg);border-color:var(--green)}.cell.v-green .w-tag{color:var(--green)}
.v-yellow{border-left-color:var(--yellow)}.cell.v-yellow{background:var(--yellow-bg);border-color:var(--yellow)}.cell.v-yellow .w-tag{color:var(--yellow)}
.v-mut{border-left-color:var(--mut)}.cell.v-mut{background:var(--mut-bg);border-color:var(--mut)}.cell.v-mut .w-tag{color:var(--mut)}
.v-het{border-left-color:var(--het)}.cell.v-het{background:var(--het-bg);border-color:var(--het)}.cell.v-het .w-tag{color:var(--het)}
.v-broken{border-left-color:var(--broken)}.cell.v-broken{background:var(--broken-bg);border-color:var(--broken)}.cell.v-broken .w-tag{color:var(--broken)}
.v-mix{border-left-color:var(--mix)}.cell.v-mix{background:var(--mix-bg);border-color:var(--mix)}.cell.v-mix .w-tag{color:var(--mix)}
.v-empty{border-left-color:var(--empty)}.cell.v-empty{background:var(--empty-bg);border-color:var(--empty)}.cell.v-empty .w-tag{color:var(--empty)}
.v-unk{border-left-color:var(--unk)}.cell.v-unk{background:var(--unk-bg)}.cell.v-unk .w-tag{color:var(--unk)}
.v-none{border-left-color:var(--none)}.cell.v-none{background:var(--none-bg)}.cell.v-none .w-tag{color:var(--none)}
.tile.v-green{border-left-color:var(--green)}.tile.v-yellow{border-left-color:var(--yellow)}.tile.v-mut{border-left-color:var(--mut)}.tile.v-het{border-left-color:var(--het)}.tile.v-broken{border-left-color:var(--broken)}.tile.v-mix{border-left-color:var(--mix)}.tile.v-empty{border-left-color:var(--empty)}.tile.v-unk{border-left-color:var(--unk)}.tile.v-none{border-left-color:var(--none)}
.lg.v-green i{background:var(--green)}.lg.v-yellow i{background:var(--yellow)}.lg.v-mut i{background:var(--mut)}.lg.v-het i{background:var(--het)}.lg.v-broken i{background:var(--broken)}.lg.v-mix i{background:var(--mix)}.lg.v-empty i{background:var(--empty)}.lg.v-unk i{background:var(--unk)}.lg.v-none i{background:var(--none)}
.cards-wrap{margin-top:34px}
.tag-crit,.tag-bb{font-size:11px;font-weight:700;padding:1px 6px;border-radius:5px}
.tag-crit{background:var(--mut-bg);color:var(--mut)}.tag-bb{background:var(--unk-bg);color:var(--ink2)}
.cards{display:grid;grid-template-columns:repeat(auto-fill,minmax(330px,1fr));gap:12px;margin-top:14px}
.card{border:1px solid var(--line);border-radius:12px;background:var(--panel);overflow:hidden;display:flex;flex-direction:column}
.card.f-repick{border-color:var(--mut)}
.chead{display:flex;align-items:center;gap:8px;padding:10px 12px;border-bottom:1px solid var(--line);flex-wrap:wrap}
.cwell{font-weight:700;font-size:14px}
.cconstruct{font-size:11px;color:var(--ink2);margin-left:auto}
.fnbadge{font-size:10px;font-weight:800;padding:2px 7px;border-radius:6px;letter-spacing:.03em}
.fnbadge.f-pick{background:var(--green-bg);color:var(--green)}.fnbadge.f-repick{background:var(--mut-bg);color:var(--mut)}.fnbadge.f-none{background:var(--unk-bg);color:var(--ink2)}
.cbody{padding:6px 12px;display:flex;flex-direction:column}
.frow{display:flex;gap:10px;padding:7px 0;border-bottom:1px dashed var(--line);align-items:baseline}
.frow:last-child{border-bottom:none}
.feat{font-size:12px;font-weight:600;min-width:96px;display:flex;flex-direction:column;line-height:1.3}
.feat em{font-style:normal;font-size:9.5px;font-weight:600;letter-spacing:.05em;text-transform:uppercase;margin-top:1px}
.frow.crit .feat em{color:var(--mut)}.frow.bb .feat em{color:var(--ink2)}
.chgs{display:flex;flex-wrap:wrap;gap:5px}
.chg{font-family:ui-monospace,monospace;font-size:11.5px;background:var(--bg);border:1px solid var(--line);border-radius:6px;padding:2px 7px;white-space:nowrap}
.chg b{color:var(--ink)}.frow.crit .chg b{color:var(--mut)}
.chg .frac{color:var(--ink2);font-size:10px}
.cfoot{margin-top:auto;padding:9px 12px;font-size:12px;color:var(--ink2);background:var(--bg);border-top:1px solid var(--line)}
.card.f-pick .cfoot b{color:var(--green)}.card.f-repick .cfoot b{color:var(--mut)}
.tbl-wrap{margin-top:34px}
.tbl-wrap h2{font-size:18px;margin:0 0 4px}.muted{color:var(--ink2);font-weight:400;font-size:14px}
.note{color:var(--ink2);font-size:13px;margin:.3em 0 1em;max-width:74ch}
.scroll{overflow-x:auto;border:1px solid var(--line);border-radius:12px;background:var(--panel)}
table{border-collapse:collapse;width:100%;font-size:13px;min-width:680px}
table.narrow{min-width:360px}
th{text-align:left;font-size:11px;letter-spacing:.06em;text-transform:uppercase;color:var(--ink2);padding:10px 12px;border-bottom:1px solid var(--line);font-weight:600;white-space:nowrap}
td{padding:9px 12px;border-bottom:1px solid var(--line)}
tr:last-child td{border-bottom:none}
td.num{text-align:right}td.small{font-size:12px;color:var(--ink2)}
.pill{display:inline-block;padding:2px 9px;border-radius:999px;font-size:11px;font-weight:700;letter-spacing:.03em;background:var(--unk-bg);color:var(--unk)}
.pill.v-green{background:var(--green-bg);color:var(--green)}.pill.v-yellow{background:var(--yellow-bg);color:var(--yellow)}.pill.v-mut{background:var(--mut-bg);color:var(--mut)}.pill.v-het{background:var(--het-bg);color:var(--het)}.pill.v-broken{background:var(--broken-bg);color:var(--broken)}.pill.v-mix{background:var(--mix-bg);color:var(--mix)}.pill.v-empty{background:var(--empty-bg);color:var(--empty)}
.ft{margin-top:40px;padding-top:16px;border-top:1px solid var(--line);font-size:11.5px;color:var(--ink2)}
.cell{cursor:pointer}
.pm-modal{display:none;position:fixed;inset:0;background:rgba(10,14,20,.5);z-index:50;padding:24px;align-items:flex-start;justify-content:center}
.pm-modal.open{display:flex}
.pm-card{background:var(--panel);border:1px solid var(--line);border-radius:14px;max-width:880px;width:100%;margin-top:6vh;max-height:82vh;overflow:auto;padding:18px 22px 22px;box-shadow:0 24px 70px rgba(0,0,0,.35)}
.pm-h{display:flex;align-items:center;gap:10px;flex-wrap:wrap}
.pm-h .pm-well{font-size:21px;font-weight:800}
.pm-x{margin-left:auto;cursor:pointer;color:var(--ink2);font-size:24px;line-height:1;padding:0 4px}
.pm-badge{font-size:11px;font-weight:800;padding:3px 10px;border-radius:6px;letter-spacing:.03em}
.pm-badge.good{background:var(--green-bg);color:var(--green)}.pm-badge.bad{background:var(--mut-bg);color:var(--mut)}.pm-badge.unknown{background:var(--unk-bg);color:var(--ink2)}
.pm-sample{font-size:12px;color:var(--ink2);font-family:ui-monospace,monospace;word-break:break-all;margin:6px 0 14px}
.health{border:1px solid var(--line);border-left-width:4px;border-radius:8px;padding:12px 14px;margin:14px 0 18px;background:var(--bg)}
.health.ok{border-left-color:var(--green)}
.health.warn{border-left-color:var(--yellow)}
.health.bad{border-left-color:var(--mut)}
.health .h-msg{font-weight:700;font-size:14px;margin-bottom:8px}
.health.ok .h-msg{color:var(--green)}
.health.warn .h-msg{color:var(--yellow)}
.health.bad .h-msg{color:var(--mut)}
.health .h-nums{display:flex;flex-wrap:wrap;gap:6px 22px;font-size:12.5px;color:var(--ink2)}
.health .h-nums b{color:var(--ink);font-variant-numeric:tabular-nums}
.health .h-nums i{font-style:normal;opacity:.65}
.health .h-note{margin-top:9px;font-size:12px;color:var(--ink2);line-height:1.5;max-width:78ch}
.pm-stats{display:flex;gap:8px 20px;flex-wrap:wrap;font-size:13px;color:var(--ink2);margin-bottom:12px}
.pm-stats b{color:var(--ink);font-variant-numeric:tabular-nums}
.pm-note{font-size:13px;background:var(--bg);border:1px solid var(--line);border-radius:8px;padding:8px 11px;margin:8px 0}
.pm-note b{color:var(--ink)}
.pm-sub{font-size:12px;letter-spacing:.06em;text-transform:uppercase;color:var(--ink2);margin:16px 0 6px}
table.pm-t{border-collapse:collapse;width:100%;font-size:12.5px}
table.pm-t th{text-align:left;font-size:10px;letter-spacing:.05em;text-transform:uppercase;color:var(--ink2);padding:5px 8px;border-bottom:1px solid var(--line)}
table.pm-t td{padding:5px 8px;border-bottom:1px solid var(--line);vertical-align:top}
table.pm-t tr.crit td{background:var(--mut-bg)}
table.pm-t tr.muted td{opacity:.55}
td.pm-1strand{color:var(--mut);font-weight:700}
tr.pm-crow{cursor:pointer}tr.pm-crow:hover td{background:var(--line)}
.imp-ok{color:var(--green);font-weight:600}.imp-mis{color:var(--het);font-weight:700}
.imp-unc{color:#8a8a8a;font-weight:600;font-style:italic}
.imp-stop{color:#fff;background:var(--mut);font-weight:700;padding:1px 6px;border-radius:4px}
.pm-ctxwrap{padding:8px 10px;background:var(--bg);border-radius:6px;margin:2px 0}
.pm-seqlbl{font-size:11px;color:var(--ink2);margin-bottom:6px}
.pm-seq{font-family:ui-monospace,Menlo,Consolas,monospace;font-size:13px;line-height:1.35;white-space:pre;letter-spacing:1px;overflow-x:auto}
.pm-dl{display:inline-flex;align-items:center;gap:6px;cursor:pointer;font-size:12px;font-weight:700;color:var(--accent);background:var(--bg);border:1px solid var(--accent);border-radius:8px;padding:5px 11px;margin:2px 0}
.pm-dlnote{font-size:11px;color:var(--ink2);margin:4px 0 2px;max-width:78ch}
.pm-seqtag{color:var(--ink2);font-size:11px;letter-spacing:0}
.pm-caret{color:var(--mut);margin-top:-2px}
.sb-mm{color:var(--het);font-weight:700}
.sb-chg{color:#fff;background:var(--mut);font-weight:700;border-radius:2px}
.pm-impline{margin-top:8px;font-size:12.5px}
.pm-dim{color:var(--ink2)}
.pm-legend{margin-top:8px;font-size:12px;color:var(--ink2);background:var(--bg);border-radius:6px;padding:8px 12px}
.pm-legend ul{margin:4px 0;padding-left:18px}.pm-legend li{margin:2px 0}
.pm-crit{color:var(--mut);font-weight:700;font-size:10px}.pm-bb{color:var(--ink2);font-size:10px}
.pm-cands{display:flex;flex-wrap:wrap;gap:6px}
.pm-cand{font-family:ui-monospace,monospace;font-size:11.5px;background:var(--bg);border:1px solid var(--line);border-radius:6px;padding:2px 8px}
.pm-track{position:relative;height:26px;border:1px solid var(--line);border-radius:6px;overflow:hidden;background:var(--mut-bg)}
.pm-bin{position:absolute;top:0;bottom:0}
.pm-tick{position:absolute;top:-2px;width:2px;height:30px;background:var(--ink)}
.pm-tick.crit{background:var(--mut)}.pm-tick.bb{background:var(--ink2);opacity:.6}
.pm-tick2{position:absolute;top:-2px;height:30px;border-radius:1px}
/* Feature lane: the map's regions drawn on the SAME x-axis as the coverage bar, so a
   gap is read as "the barcode is missing" without mapping a chip back onto the bar. */
.pm-flane{position:relative;height:34px;margin-top:2px}
.pm-freg{position:absolute;top:0;height:32px;border-radius:4px;border:1px solid var(--line);
  background:var(--bg);overflow:hidden;display:flex;align-items:center;justify-content:center;
  font-size:10px;line-height:1.1;padding:0 3px;text-align:center;box-sizing:border-box;cursor:default}
.pm-freg.ok{background:var(--green-bg);border-color:var(--green);color:var(--green)}
.pm-freg.part{background:var(--yellow-bg);border-color:var(--yellow);color:var(--yellow)}
.pm-freg.gap{background:var(--mut-bg);border-color:var(--mut);color:var(--mut);font-weight:700}
.pm-freg.crit{outline:1px solid currentColor;outline-offset:-3px}
.pm-freg span{overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:100%}
.pm-flist{display:flex;flex-wrap:wrap;gap:5px}
.pm-fchip{font-size:11px;border-radius:6px;padding:2px 8px;border:1px solid var(--line)}
.pm-fchip.ok{background:var(--green-bg);color:var(--green)}
.pm-fchip.part{background:var(--yellow-bg);color:var(--yellow)}
.pm-fchip.gap{background:var(--mut-bg);color:var(--mut)}
.pm-fchip.crit{font-weight:700;outline:1px solid currentColor}
.mono{font-family:ui-monospace,monospace}
@media(max-width:560px){.hd h1{font-size:27px}article{padding:28px 16px 60px}.cards{grid-template-columns:1fr}}
"""

_HTML_JS = r"""
function pmEsc(s){return String(s==null?'':s).replace(/[&<>]/g,function(c){return{'&':'&amp;','<':'&lt;','>':'&gt;'}[c];});}
function pmShow(k){
  var w=WELLS[k]; if(!w) return;
  var h='<div class="pm-h"><span class="pm-well">'+pmEsc(w.well)+'</span>'+
        '<span class="pm-badge '+w.state+'">'+pmEsc(w.head)+'</span>'+
        '<span class="pm-x" onclick="pmClose()" title="close">&times;</span></div>'+
        '<div class="pm-sample">'+pmEsc(w.sample)+'</div>';
  h+='<div class="pm-stats">';
  if(w.called){h+='<span>construct <b>'+pmEsc(w.called)+'</b></span>';}
  h+='<span>breadth <b>'+w.breadth+'%</b></span><span>depth <b>'+w.depth+'x</b></span>';
  if(w.support){h+='<span>unique reads <b>'+w.support+'</b></span>';}
  h+='</div>';
  if(w.reason){h+='<div class="pm-note"><b>Note:</b> '+pmEsc(w.reason)+'</div>';}
  if(w.corrected_gb){h+='<div class="pm-dlnote"><span class="pm-dl" onclick="pmDl('+"'"+pmEsc(k)+"'"+')">&#11015; Download corrected .gb</span> '+
    'consensus of your reads on the called map; grey &lsquo;unverified&rsquo; regions had no coverage and are copied from the map.</div>';}
  if(w.runner){h+='<div class="pm-note">also present: <b>'+pmEsc(w.runner)+'</b> (possible contamination)</div>';}
  if(w.background&&w.background.length){h+=w.background.map(function(b){
    return '<div class="pm-note"><b>Also in this well:</b> '+pmEsc(b[0])+' at <b>'+b[1]+'x</b>'+
      ' &mdash; it runs <b>'+b[2]+'x</b> in the wells where it is the real clone, so this'+
      ' is read as plate-wide background rather than your pick. The call above was made'+
      ' with it set aside.</div>';}).join('');}
  if(w.deletion){h+='<div class="pm-note"><b>Deletion:</b> ~'+w.deletion.bp+' bp missing at '+pmEsc(w.deletion.span)+' ('+pmEsc(w.deletion.feat)+')</div>';}
  if(w.track&&w.track.length){
    var bw=100/w.track.length;
    // Track 1 — COVERAGE: what was sequenced vs not sequenced
    h+='<div class="pm-sub">Coverage across '+pmEsc(w.track_label||'the map')+' ('+w.track_len+' bp) &mdash; <span style="color:var(--green)">green = sequenced</span>, <span style="color:var(--mut)">red = not sequenced (no reads here)</span></div>';
    h+='<div class="pm-track">';
    for(var t=0;t<w.track.length;t++){var cv=w.track[t];
      var col=cv>=50?'var(--green)':(cv>=15?'var(--yellow)':'var(--mut)');
      h+='<span class="pm-bin" style="left:'+(t*bw)+'%;width:'+(bw+0.25)+'%;background:'+col+';opacity:'+(cv>0?(0.4+0.6*cv/100):1)+'"></span>';
    }
    h+='</div>';
    // Region lane, drawn on the SAME axis as the bar above: read a gap straight off as
    // "which part of the construct is missing" rather than mapping a chip back onto it.
    if(w.feats&&w.feats.length){
      h+='<div class="pm-flane">';
      for(var gi=0;gi<w.feats.length;gi++){var gf=w.feats[gi];
        var g0=Math.floor(gf[0]/100*w.track.length),g1=Math.ceil(gf[1]/100*w.track.length),gs=0,gc=0;
        for(var gb=g0;gb<g1&&gb<w.track.length;gb++){gs+=w.track[gb];gc++;}
        var gp=gc?Math.round(gs/gc):0, gcls=gp>=80?'ok':(gp>=30?'part':'gap');
        var wd=Math.max(0.8,gf[1]-gf[0]);
        h+='<span class="pm-freg '+gcls+(gf[3]==='critical'?' crit':'')+'" style="left:'+gf[0]+'%;width:'+wd+'%" title="'+pmEsc(gf[2])+' — '+gp+'% sequenced ('+Math.round(gf[0]/100*w.track_len)+'-'+Math.round(gf[1]/100*w.track_len)+' bp)">'
          +(wd>=6?'<span>'+pmEsc(gf[2])+'</span>':'')+'</span>';
      }
      h+='</div>';
      h+='<div class="pm-sub" style="margin-top:8px">Regions above are drawn on the same scale as the bar &mdash; a red block is a part of the construct no read covered, so it could not be checked</div>';
    }
    // Track 2 — AGREEMENT: where the sequenced reads match the map vs differ
    h+='<div class="pm-sub">Agreement with the map &mdash; <span style="color:var(--green)">green = matches</span>; ticks: <span style="color:var(--mut)">red = this clone changed</span>, <span style="color:var(--accent)">blue = map is stale (shared)</span>, <span style="color:var(--yellow)">amber = low-level</span></div>';
    h+='<div class="pm-track">';
    for(var t2=0;t2<w.track.length;t2++){var cv2=w.track[t2];
      var mcol=cv2>0?'var(--green)':'var(--line)';
      h+='<span class="pm-bin" style="left:'+(t2*bw)+'%;width:'+(bw+0.25)+'%;background:'+mcol+';opacity:'+(cv2>0?0.85:1)+'"></span>';
    }
    if(w.marks){for(var mi=0;mi<w.marks.length;mi++){var mk=w.marks[mi];var kd=mk[2];
      var tcol=(kd==='mutation'||kd==='heterogeneous')?'var(--mut)':(kd==='systematic')?'var(--accent)':'var(--yellow)';
      var tz=(kd==='mutation'||kd==='heterogeneous')?'3px':'2px';
      h+='<span class="pm-tick2" style="left:'+mk[0]+'%;width:'+tz+';background:'+tcol+'" title="'+pmEsc(kd)+' at '+mk[0]+'% ('+mk[1]+')"></span>';}}
    h+='</div>';
    if(w.feats&&w.feats.length){
      h+='<div class="pm-sub">Is each region present?</div><div class="pm-flist">';
      for(var fi=0;fi<w.feats.length;fi++){var ff=w.feats[fi];
        var b0=Math.floor(ff[0]/100*w.track.length),b1=Math.ceil(ff[1]/100*w.track.length),su=0,cn=0;
        for(var bb=b0;bb<b1&&bb<w.track.length;bb++){su+=w.track[bb];cn++;}
        var cvp=cn?Math.round(su/cn):0, cls=cvp>=80?'ok':(cvp>=30?'part':'gap');
        h+='<span class="pm-fchip '+cls+(ff[3]==='critical'?' crit':'')+'">'+pmEsc(ff[2])+' <b>'+cvp+'%</b></span>';
      }
      h+='</div>';
    }
  }
  if(w.changes&&w.changes.length){
    h+='<div class="pm-sub">Exact changes vs the map ('+w.changes.length+') &mdash; click a row to see it in sequence context</div>'+
       '<div style="overflow-x:auto"><table class="pm-t"><thead><tr><th>pos</th><th>change</th><th>reads</th><th>strands</th><th>region</th><th>type</th><th>impact</th></tr></thead><tbody>';
    for(var i=0;i<w.changes.length;i++){var c=w.changes[i];
      var reg=(c.cat==='critical')?'<span class="pm-crit">critical</span>':'<span class="pm-bb">backbone</span>';
      var rc=(c.kind==='mutation'||c.kind==='heterogeneous')?(c.cat==='critical'?'crit':''):'muted';
      // strand balance: a real variant is on both strands; one-sided = prep/oxidation
      var sm=/^(\d+)\+\/(\d+)-$/.exec(c.strand||''); var sflag='';
      if(sm){var fw=+sm[1],rv=+sm[2],tot=fw+rv; if(tot>=6 && Math.min(fw,rv)/tot<0.10) sflag=' pm-1strand';}
      // protein-level impact, if this change sits in a coding feature
      var imp='&mdash;', icls='';
      if(c.impact){var ik=c.impact[0];
        icls=(ik==='stop'||ik==='frameshift')?'imp-stop':(ik==='missense'||ik==='readthrough'||ik==='inframe_indel')?'imp-mis':(ik==='uncertain')?'imp-unc':'imp-ok';
        var ilbl=(ik==='stop')?'STOP':(ik==='frameshift')?'FRAMESHIFT':(ik==='inframe_indel')?'in-frame indel':(ik==='uncertain')?'frame?':ik;
        imp='<span class="'+icls+'" title="'+pmEsc(c.impact[1])+'">'+pmEsc(ilbl)+'</span>';}
      h+='<tr class="pm-crow '+rc+'" onclick="pmTog(this)"><td class="mono">'+c.pos+' &#9662;</td>'+
         '<td class="mono"><b>'+pmEsc(c.chg)+'</b></td><td class="mono">'+pmEsc(c.reads)+' ('+c.pct+'%)</td>'+
         '<td class="mono'+sflag+'" title="alt reads forward/reverse — balanced = real, one-sided = artifact">'+pmEsc(c.strand||'—')+'</td>'+
         '<td>'+pmEsc(c.feat)+' '+reg+'</td><td>'+pmEsc(c.kind)+(c.note?' &middot; '+pmEsc(c.note):'')+'</td><td>'+imp+'</td></tr>';
      // hidden context row: the local map sequence vs what the reads said
      h+='<tr class="pm-ctxrow" style="display:none"><td colspan="7">'+pmCtx(c)+'</td></tr>';
    }
    h+='</tbody></table></div>';
    // explain the greyed-out rows: they are NOT held against the clone, and why
    var kinds={};for(var ki=0;ki<w.changes.length;ki++){kinds[w.changes[ki].kind]=1;}
    var leg=[];
    if(kinds.artifact)   leg.push('<b>artifact</b> = found on essentially one DNA strand (e.g. 8-oxoG oxidation / prep) &mdash; not a real change');
    if(kinds.minor)      leg.push('<b>minor</b> = a low-level subpopulation (&lt;30% of reads) &mdash; background, not the main sequence');
    if(kinds.systematic) leg.push('<b>systematic</b> = the same change appears across the whole plate &mdash; the map (.gb) is stale, not your clone');
    if(leg.length){
      h+='<div class="pm-legend"><b>Greyed-out rows are not counted against this clone.</b><ul><li>'
         +leg.join('</li><li>')+'</li></ul>Only <b>mutation</b> and <b>heterogeneous</b> (dark rows) affect the verdict.</div>';
    }
  } else if(w.called){
    h+='<div class="pm-note">No disqualifying changes &mdash; the consensus matches the map.</div>';
  }
  if(w.candidates&&w.candidates.length){
    h+='<div class="pm-sub">Best-matching maps (how completely the reads cover each)</div><div class="pm-cands">';
    for(var j=0;j<w.candidates.length;j++){h+='<span class="pm-cand">'+pmEsc(w.candidates[j][0])+' <b>'+w.candidates[j][1]+'%</b></span>';}
    h+='</div>';
  }
  document.getElementById('pm-card').innerHTML=h;
  document.getElementById('pm-modal').classList.add('open');
}
function pmTog(row){var d=row.nextElementSibling;
  if(d&&d.className==='pm-ctxrow'){d.style.display=(d.style.display==='none')?'':'none';}}
function pmCtx(c){
  if(!c.ctx_ref){return '<div class="pm-ctxwrap"><span class="pm-dim">no sequence context available</span></div>';}
  var ref=c.ctx_ref, obs=c.ctx_obs||'', k=c.ctx_i, start=c.ctx_start;
  var mline='', oline='';
  for(var i=0;i<ref.length;i++){
    var rb=ref[i], ob=obs[i]||'.';
    var cls=(i===k)?'sb-chg':(ob!=='.'&&ob!==rb?'sb-mm':'');
    mline+='<span class="'+((i===k)?'sb-chg':'')+'">'+rb+'</span>';
    oline+='<span class="'+cls+'">'+ob+'</span>';
  }
  var pos='&nbsp;'.repeat(k>0?k:0)+'&#9650;';   // caret under the changed base
  var imp='';
  if(c.impact){var ik=c.impact[0];
    var ic=(ik==='stop')?'imp-stop':(ik==='missense'||ik==='readthrough')?'imp-mis':(ik==='uncertain')?'imp-unc':'imp-ok';
    imp='<div class="pm-impline '+ic+'"><b>protein impact (predicted):</b> '+pmEsc(c.impact[1])+'</div>';}
  return '<div class="pm-ctxwrap">'+
    '<div class="pm-seqlbl">around position '+start+' &mdash; upper = <b>map</b>, lower = <b>reads</b> ('
    +'<span class="sb-mm">orange</span> = mismatch, <span class="sb-chg">red</span> = this change; <span class="pm-dim">.</span> = not sequenced)</div>'+
    '<div class="pm-seq"><span class="pm-seqtag">map&nbsp;&nbsp;</span>'+mline+'</div>'+
    '<div class="pm-seq"><span class="pm-seqtag">reads</span>'+oline+'</div>'+
    '<div class="pm-seq pm-caret"><span class="pm-seqtag">&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</span>'+pos+'</div>'+
    imp+'</div>';
}
function pmClose(){document.getElementById('pm-modal').classList.remove('open');}
function pmDl(k){var w=WELLS[k]; if(!w||!w.corrected_gb) return;
  // fully offline: build the .gb text into a Blob and trigger a download
  var blob=new Blob([w.corrected_gb],{type:'chemical/seq-na-genbank'});
  var url=URL.createObjectURL(blob);
  var a=document.createElement('a');
  a.href=url; a.download=(w.well+'_'+(w.called||'construct')+'_corrected.gb').replace(/[^\w.\-]+/g,'_');
  document.body.appendChild(a); a.click(); document.body.removeChild(a);
  setTimeout(function(){URL.revokeObjectURL(url);},1500);
}
document.addEventListener('keydown',function(e){if(e.key==='Escape')pmClose();});
"""


def write_html(results, path):
    esc = _html.escape
    counts = Counter(r.verdict for r in results)
    called = [r for r in results if r.called]
    decs = [_decision(r) for r in results]
    n_good = sum(1 for s, _, _ in decs if s == "good")
    n_bad = sum(1 for s, _, _ in decs if s == "bad")
    n_unknown = sum(1 for s, _, _ in decs if s == "unknown")
    n_perfect = sum(1 for _, _, p in decs if p)
    run_label = Path(path).parent.parent.name or "pickme run"

    # summary tiles — three states, nothing else
    tiles = (
        f'<div class="tile d-good"><div class="tnum">{n_good}</div>'
        f'<div class="tlbl">Use</div>'
        f'<div class="thuman">pickable · {n_perfect} flawless</div></div>'
        f'<div class="tile d-bad"><div class="tnum">{n_bad}</div>'
        f'<div class="tlbl">Don&rsquo;t use</div>'
        f'<div class="thuman">re-pick / re-streak</div></div>'
        f'<div class="tile d-unknown"><div class="tnum">{n_unknown}</div>'
        f'<div class="tlbl">Unknown</div>'
        f'<div class="thuman">no confident match</div></div>')

    # plate grids
    by_plate = defaultdict(list)
    for r in results:
        by_plate[r.plate].append(r)

    def cell(r):
        state, mark, perfect = _decision(r)
        br = r.breadth * 100
        call = _short(r.called)
        detail = (_func_call(r)[2].replace("&rarr;", ">")
                  .replace("&mdash;", "-").replace("&ndash;", "-"))
        head = {"good": "USE — pickable", "bad": "DON'T USE — re-pick",
                "unknown": "UNKNOWN — no confident match"}[state]
        # full status shown on hover
        tip = f"{r.well} · {r.sample}\n{head}"
        if call:
            tip += f"\ncalled: {call}\n{br:.0f}% breadth · {r.mean_depth:.0f}x depth"
            if r.member_support:
                tip += f"\n{r.member_support} unique reads name it"
        else:
            tip += f"\nbest breadth {br:.0f}%"
            if r.reason:
                tip += f"\nwhy: {r.reason}"
        if r.candidates:
            tip += "\nbest matches: " + ", ".join(f"{_short(n)} {b*100:.0f}%"
                                                   for n, b in r.candidates[:3])
        if r.n_mut:
            tip += f"\n{r.n_mut} mutation(s)"
        if r.n_het:
            tip += f"\n{r.n_het} heterogeneous position(s)"
        if r.runner_up:
            tip += f"\nalso present: {_short(r.runner_up)}"
        tip += f"\ndetail: {detail}"
        markcls = "w-mark star" if perfect else "w-mark"
        inner = f'<span class="w-well">{esc(r.well)}</span>'
        inner += (f'<span class="w-call">{esc(call)}</span>' if call
                  else f'<span class="w-call w-dim">{br:.0f}%</span>')
        key = f"{r.plate}|{r.well}"
        return (f'<div class="cell d-{state}" title="{esc(tip)}" '
                f'data-k="{esc(key)}" onclick="pmShow(this.dataset.k)">'
                f'<span class="{markcls}">{mark}</span>{inner}</div>')

    plate_html = ""
    for plate in sorted(by_plate):
        wells = by_plate[plate]
        used_cols = sorted({int(WELL_RE.match(r.well).group(2))
                            for r in wells if WELL_RE.match(r.well)})
        by_well = {r.well: r for r in wells}
        g = [f'<div class="grid" style="--cols:{len(used_cols)}"><div class="ghead corner"></div>']
        for c in used_cols:
            g.append(f'<div class="ghead">{c}</div>')
        for row in "ABCDEFGH":
            if not any(r.well.startswith(row) for r in wells):
                continue
            g.append(f'<div class="ghead rlab">{row}</div>')
            for c in used_cols:
                wk = f"{row}{c}"
                g.append(cell(by_well[wk]) if wk in by_well
                         else '<div class="cell empty"></div>')
        g.append("</div>")
        plate_html += (f'<section class="plate"><h3>{esc(plate)}'
                       f'<span class="pcount">{len(wells)} wells</span></h3>'
                       f'{"".join(g)}</section>')

    # per-well diff cards (real problems first)
    flagged = [r for r in results
               if r.verdict in ("RED_MUT", "RED_HET", "RED_BROKEN", "RED_MIX")]
    flagged.sort(key=lambda r: (_func_call(r)[0] != "repick", r.well))
    cards = ""
    for r in flagged:
        fstatus, fbadge, fdetail = _func_call(r)
        tag, cls, human = _HTML_VMAP[r.verdict]
        groups = defaultdict(list)
        for d in r.differences:
            if d.kind not in ("mutation", "heterogeneous"):
                continue
            groups[(d.features or ["unannotated region"])[0]].append(d)
        rows = ""
        for feat in sorted(groups, key=lambda f: _feat_cat([f]) != "critical"):
            cat = _feat_cat([feat])
            chips = ""
            for d in sorted(groups[feat], key=lambda x: x.pos):
                het = d.kind == "heterogeneous"
                frac = f' <span class="frac">{d.fraction * 100:.0f}%{" ·2 alleles" if het else ""}</span>'
                if d.indel_kind:             # indel: show +GT/-A and the frameshift tag
                    tag = " frameshift" if "frameshift" in d.note else (
                        " in-frame" if "in-frame" in d.note else "")
                    body = f'{esc(d.observed)}<span class="frac">{esc(tag)}</span>'
                else:
                    body = f'{d.expected}&rarr;{d.observed}'
                chips += (f'<span class="chg">{d.pos + 1} '
                          f'<b>{body}</b>{frac}</span>')
            rows += (f'<div class="frow {"crit" if cat == "critical" else "bb"}">'
                     f'<span class="feat">{esc(feat.split(" (")[0])}<em>{cat}</em></span>'
                     f'<span class="chgs">{chips}</span></div>')
        if r.deletion:
            s, e, bp = r.deletion
            df = ", ".join(x.split(" (")[0] for x in (r.deletion_features or [])) or "&mdash;"
            rows = (f'<div class="frow crit"><span class="feat">deletion<em>structural</em></span>'
                    f'<span class="chgs"><span class="chg">~{bp} bp missing at '
                    f'{s + 1}&ndash;{e} <b>({esc(df)})</b></span></span></div>' + rows)
        if r.verdict == "RED_MIX" and r.runner_up:
            rows = (f'<div class="frow crit"><span class="feat">second construct'
                    f'<em>contamination</em></span><span class="chgs"><span class="chg">'
                    f'also carries <b>{esc(_short(r.runner_up))}</b></span></span></div>' + rows)
        foot = ("Backbone-only &mdash; <b>usable</b> for your construct."
                if fstatus == "pick"
                else "Hits a region that matters &mdash; <b>re-pick</b>.")
        cards += (f'<div class="card f-{fstatus}"><div class="chead">'
                  f'<span class="cwell mono">{esc(r.well)}</span>'
                  f'<span class="pill {cls}">{tag}</span>'
                  f'<span class="fnbadge f-{fstatus}">{fbadge}</span>'
                  f'<span class="cconstruct mono">{esc(_short(r.called) or "—")}</span></div>'
                  f'<div class="cbody">{rows or "<div class=frow><span class=chgs>no disqualifying changes</span></div>"}</div>'
                  f'<div class="cfoot">{foot}</div></div>')

    # called-wells table
    def flagtxt(r):
        p = []
        if r.n_mut:
            p.append(f"{r.n_mut} mut")
        if r.n_het:
            p.append(f"{r.n_het} het")
        if r.n_systematic:
            p.append(f"{r.n_systematic} sys")
        if r.runner_up:
            p.append(f"+{_short(r.runner_up)}")
        return ", ".join(p) or "&mdash;"

    trows = ""
    for r in sorted(called, key=lambda r: (_HTML_ORDER.index(r.verdict), r.sample)):
        tag, cls, human = _HTML_VMAP[r.verdict]
        fstatus, fbadge, _ = _func_call(r)
        trows += (f'<tr><td class="mono">{esc(r.well)}</td><td>{esc(r.sample)}</td>'
                  f'<td><span class="pill {cls}">{tag}</span></td>'
                  f'<td><span class="fnbadge f-{fstatus}">{fbadge}</span></td>'
                  f'<td class="mono">{esc(_short(r.called))}</td>'
                  f'<td class="mono num">{r.breadth * 100:.0f}%</td>'
                  f'<td class="mono num">{r.mean_depth:.0f}x</td>'
                  f'<td class="mono num small">{r.member_support or "&mdash;"}</td>'
                  f'<td class="mono small">{flagtxt(r)}</td></tr>')

    # unknown by family
    fam = defaultdict(list)
    for r in results:
        if r.verdict == "RED_UNKNOWN":
            base = re.split(r"_aRY|_S\d+_", r.sample)[0]
            fam[base].append(r)
    unrows = ""
    for f in sorted(fam):
        brs = sorted(x.breadth * 100 for x in fam[f])
        why = Counter(x.reason for x in fam[f] if x.reason).most_common(1)
        why = why[0][0] if why else "&mdash;"
        cand = next((x.candidates for x in fam[f] if x.candidates), [])
        cand_s = ", ".join(f"{_short(n)} {b*100:.0f}%" for n, b in cand[:2]) or "&mdash;"
        unrows += (f'<tr><td>{esc(f)}</td><td class="mono num">{len(fam[f])}</td>'
                   f'<td class="mono num">{min(brs):.0f}&ndash;{max(brs):.0f}%</td>'
                   f'<td class="mono small">{esc(cand_s)}</td>'
                   f'<td class="small">{esc(why)}</td></tr>')

    legend = (
        '<span class="lg d-good"><i></i>Use</span>'
        '<span class="lg d-bad"><i></i>Don&rsquo;t use</span>'
        '<span class="lg d-unknown"><i></i>Unknown</span>'
        '<span class="lg">&nbsp;marks: <b>&#9733;</b> flawless &middot; '
        '<b>~</b> partial coverage &middot; <b>bb</b> backbone-only edit &middot; '
        '<b>het</b> heterogeneous &middot; <b>MUT/HET/MIX/DEL</b> = reason it&rsquo;s not usable '
        '&middot; hover any well for its full status</span>')

    n_unk = sum(len(v) for v in fam.values())

    # per-well detail for the click-to-open panel (exact changes, evidence)
    def detail_record(r):
        state, mark, perfect = _decision(r)
        changes = []
        for dff in r.differences:
            if dff.kind not in ("mutation", "heterogeneous", "minor",
                                "systematic", "artifact"):
                continue
            chg = dff.observed if dff.indel_kind else f"{dff.expected}→{dff.observed}"
            changes.append(dict(
                pos=dff.pos + 1, chg=chg,
                pct=round(dff.fraction * 100), reads=f"{dff.alt_count}/{dff.depth}",
                strand=f"{dff.alt_fwd}+/{dff.alt_rev}-",
                feat=(dff.features[0].split(" (")[0] if dff.features else "—"),
                cat=_feat_cat(dff.features), kind=dff.kind, note=dff.note or "",
                indel=dff.indel_kind or "",
                ctx_ref=dff.ctx_ref, ctx_obs=dff.ctx_obs, ctx_i=dff.ctx_i,
                ctx_start=dff.ctx_start + 1,
                impact=list(dff.impact) if dff.impact else None))
        # critical changes first, then by position
        changes.sort(key=lambda c: (c["cat"] != "critical", c["pos"]))
        rec = dict(
            well=r.well, sample=r.sample, state=state,
            head={"good": "USE", "bad": "DON'T USE", "unknown": "UNKNOWN"}[state],
            called=_short(r.called) or "", breadth=round(r.breadth * 100),
            depth=round(r.mean_depth), support=r.member_support,
            reason=r.reason or "", runner=_short(r.runner_up) or "",
            candidates=[[_short(n), round(b * 100)] for n, b in (r.candidates or [])[:4]],
            # what else is demonstrably IN this well but read as plate background:
            # [name, depth here, depth where it is the real clone]
            background=[[_short(n), d, md] for n, d, md in (r.background or [])],
            changes=changes)
        # coverage track: what was sequenced (green) vs missing (red), with features
        rec["track"] = r.cov_track
        rec["track_len"] = r.track_len
        rec["track_label"] = _short(r.track_label)
        rec["feats"] = [[s, e, lab.split(" (")[0], _feat_cat([lab])]
                        for s, e, lab in (r.track_features or [])]
        # EVERY difference as a tick on the agreement bar, tagged by kind so the
        # view can colour "your clone changed" (mutation/het) apart from "map is
        # stale" (systematic) and low-level background (minor/artifact)
        rec["marks"] = [[round(100 * dff.pos / r.track_len, 2),
                         _feat_cat(dff.features), dff.kind]
                        for dff in r.differences
                        if dff.kind in ("mutation", "heterogeneous", "minor",
                                        "systematic", "artifact") and r.track_len]
        if r.deletion:
            s, e, bp = r.deletion
            rec["deletion"] = dict(
                bp=bp, span=f"{s + 1}–{e}",
                feat=", ".join(x.split(" (")[0] for x in (r.deletion_features or [])) or "—")
        if r.corrected_gb:               # read-consensus GenBank, downloadable offline
            rec["corrected_gb"] = r.corrected_gb
        return rec

    wells_data = {f"{r.plate}|{r.well}": detail_record(r) for r in results}
    wells_json = json.dumps(wells_data, ensure_ascii=False).replace("</", "<\\/")

    # ── Run health: is this run even worth reading? ──────────────────────────────
    # A grey plate has two very different causes and the reader cannot tell them apart
    # from the wells alone: the tool being unsure, or the SEQUENCING never covering the
    # spot that decides. These are the numbers that separate them, up front.
    _tot = [r.total_reads for r in results]
    _map = [r.mapped_reads / r.total_reads for r in results if r.total_reads]
    _med_reads = int(sorted(_tot)[len(_tot) // 2]) if _tot else 0
    _med_map = (sorted(_map)[len(_map) // 2] * 100) if _map else 0.0
    _dirty = sum(1 for m in _map if m < 0.60)
    _thin = sum(1 for t in _tot if t < 3000)
    _blind = sum(1 for r in results
                 if not r.identified and r.mapped_reads >= MIN_MAPPED)
    _hcls = ("bad" if (_med_map < 60 or _dirty > len(results) / 2) else
             "warn" if (_med_reads < 5000 or _thin > len(results) / 10) else "ok")
    _hmsg = {"ok": "This run looks healthy.",
             "warn": "This run is usable but thin in places.",
             "bad": "Most of this run&rsquo;s reads are not your constructs."}[_hcls]
    health = f"""<div class="health {_hcls}">
  <div class="h-msg">{_hmsg}</div>
  <div class="h-nums">
    <span><b>{_med_reads:,}</b> reads per well <i>(median)</i></span>
    <span><b>{_med_map:.0f}%</b> of reads match a map you supplied <i>(median)</i></span>
    <span><b>{_dirty}</b>/{len(results)} wells &gt;40% foreign DNA</span>
    <span><b>{_thin}</b>/{len(results)} wells under 3,000 reads</span>
    <span><b>{_blind}</b> wells had reads but never covered the deciding spot</span>
  </div>
  <div class="h-note">A well can be well covered and still unnameable: Tn5 inserts at random,
    so a plasmid can be 90%+ sequenced while the barcode or insert junction &mdash; the only
    part that tells your constructs apart &mdash; gets no reads at all. Those wells are grey
    because the data cannot answer, not because the clone is bad.</div>
</div>"""

    body = f"""<article>
<header class="hd">
  <div class="eyebrow">pickme &middot; clone verification</div>
  <h1>{esc(run_label)}</h1>
  {health}
  <p class="sub">One question per well: can you use it? Green = yes, red = no, grey =
    can&rsquo;t tell. A small mark adds the nuance &mdash; a star for a flawless clone, or
    the reason a well isn&rsquo;t usable. Hover any well to see its full status.</p>
  <div class="tiles">{tiles}</div>
  <div class="headline">
    <div class="hl-main"><b>{n_good}</b> to use
      <span class="hl-sep">&middot;</span> <b>{n_bad}</b> don&rsquo;t use
      <span class="hl-sep">&middot;</span> <b>{n_unknown}</b> unknown</div>
    <div class="hl-sub"><b>{n_perfect}</b> of the usable wells are flawless (&#9733;); the rest
      are usable but carry a minor note &mdash; partial coverage, or an edit only in vector
      backbone (ori / marker / promoter) that doesn&rsquo;t change your construct.</div>
  </div>
</header>
<div class="legend">{legend}</div>
<div class="plates">{plate_html}</div>
<section class="cards-wrap">
  <h2>What changed, well by well</h2>
  <p class="note">Every flagged well, real problems first. Each row is one feature and the
    exact edits inside it &mdash; position, base change, and how many reads carry it.
    <span class="tag-crit">critical</span> = your construct&rsquo;s working parts;
    <span class="tag-bb">backbone</span> = vector plumbing that doesn&rsquo;t affect function.</p>
  <div class="cards">{cards}</div>
</section>
<section class="tbl-wrap">
  <h2>Called wells <span class="muted">({len(called)})</span></h2>
  <div class="scroll"><table>
    <thead><tr><th>Well</th><th>Sample</th><th>Match</th><th>Function</th>
      <th>Construct</th><th>Breadth</th><th>Depth</th><th>Unique reads</th>
      <th>Flags</th></tr></thead>
    <tbody>{trows}</tbody></table></div>
</section>
<section class="tbl-wrap">
  <h2>Unknown, by family <span class="muted">({n_unk})</span></h2>
  <p class="note">Matched no supplied map &mdash; either a construct with no reference in
    the folder, or a genuinely poor clone (reads piled on a fraction of the plasmid).</p>
  <div class="scroll"><table>
    <thead><tr><th>Family</th><th>Wells</th><th>Breadth range</th>
      <th>Best matches</th><th>Why unresolved</th></tr></thead>
    <tbody>{unrows}</tbody></table></div>
</section>
<footer class="ft">Generated by pickme &middot; click any well for its exact changes &middot;
  verdict = raw match to the map; the mark is a heuristic on which features matter.</footer>
</article>
<div id="pm-modal" class="pm-modal" onclick="if(event.target===this)pmClose()">
  <div class="pm-card" id="pm-card"></div>
</div>"""
    script = ("<script>\nconst WELLS=" + wells_json + ";\n" + _HTML_JS + "\n</script>")
    doc = ("<!doctype html>\n<html lang=\"en\">\n<head>\n<meta charset=\"utf-8\">\n"
           "<meta name=\"viewport\" content=\"width=device-width, initial-scale=1\">\n"
           f"<title>pickme &mdash; {esc(run_label)}</title>\n<style>{_HTML_CSS}</style>\n"
           f"</head>\n<body>\n{body}\n{script}\n</body>\n</html>\n")
    Path(path).write_text(doc, encoding="utf-8")


def write_xlsx(results, path):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fills = {"GREEN": "C8E6C9", "YELLOW": "FFF59D", "RED_MIX": "FFCC80",
             "RED_MUT": "EF9A9A", "RED_HET": "F48FB1", "RED_BROKEN": "E57373",
             "RED_EMPTY": "D7B899", "RED_UNKNOWN": "E0E0E0", "NO_DATA": "EEEEEE"}

    by_plate = defaultdict(dict)
    for r in results:
        by_plate[r.plate][r.well] = r
    best = {}
    for r in results:
        if r.verdict == "GREEN" and r.called:
            if r.called not in best or r.mean_depth > best[r.called].mean_depth:
                best[r.called] = r
    gold = Side(style="thick", color="FFB300")
    gborder = Border(left=gold, right=gold, top=gold, bottom=gold)

    # ── clean copy-paste grid: intended construct name + one of three colours,
    # matching the bench Excel sheet (green = use, red = don't, grey = can't tell) ──
    thin = Side(style="thin", color="B0B4BA")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    state_fill = {"good": "C8E6C9", "bad": "EF9A9A", "unknown": "E0E0E0"}
    hdr = Font(bold=True)
    for plate in sorted(by_plate):
        ws = wb.create_sheet(plate[:31])
        ws.cell(row=1, column=1, value=plate).font = hdr
        for col in range(1, 13):
            c = ws.cell(row=2, column=col + 1, value=col)
            c.alignment, c.font = center, hdr
        for ri, letter in enumerate("ABCDEFGH"):
            c = ws.cell(row=ri + 3, column=1, value=letter)
            c.alignment, c.font = center, hdr
        for well, r in by_plate[plate].items():
            m = WELL_RE.fullmatch(well)
            if not m:
                continue
            row = ord(m.group(1)) - ord("A") + 3
            col = int(m.group(2)) + 1
            if not (3 <= row <= 10 and 2 <= col <= 13):
                continue
            state, mark, perfect = _decision(r)
            name = intended_label(r.sample) + (" ★" if perfect else "")
            cell = ws.cell(row=row, column=col, value=name)
            cell.fill = PatternFill("solid", fgColor=state_fill.get(state, "E0E0E0"))
            cell.alignment = center
            cell.font = Font(size=9)
            cell.border = box
        ws.column_dimensions["A"].width = 4
        for col in range(2, 14):
            ws.column_dimensions[get_column_letter(col)].width = 18
        for row in range(3, 11):
            ws.row_dimensions[row].height = 26
        # tiny legend under the grid
        ws.cell(row=12, column=1, value="green = use").fill = \
            PatternFill("solid", fgColor="C8E6C9")
        ws.cell(row=12, column=2, value="red = don't use").fill = \
            PatternFill("solid", fgColor="EF9A9A")
        ws.cell(row=12, column=3, value="grey = can't tell").fill = \
            PatternFill("solid", fgColor="E0E0E0")
        ws.cell(row=12, column=4, value="★ = reference-grade")

    for plate in sorted(by_plate):
        ws = wb.create_sheet((plate[:24] + " detail"))
        for col in range(1, 13):
            ws.cell(row=1, column=col + 1, value=col).alignment = center
        for ri, letter in enumerate("ABCDEFGH"):
            ws.cell(row=ri + 2, column=1, value=letter).alignment = center
        for well, r in by_plate[plate].items():
            m = WELL_RE.fullmatch(well)
            if not m:
                continue
            row = ord(m.group(1)) - ord("A") + 2
            col = int(m.group(2)) + 1
            if not (2 <= row <= 9 and 2 <= col <= 13):
                continue
            tag = {"GREEN": "MATCH", "YELLOW": "CHECK", "RED_MUT": "MUT",
                   "RED_HET": "HET", "RED_MIX": "MIX", "RED_BROKEN": "DEL",
                   "RED_EMPTY": "EMPTY", "RED_UNKNOWN": "UNK", "NO_DATA": "—"}[r.verdict]
            txt = f"{r.called or '?'}\n{tag}"
            if r.called:
                txt += f"\n{r.mean_depth:.0f}x {r.breadth:.0%}"
                if r.n_mut:
                    txt += f"\n{r.n_mut} mut"
                elif r.n_het:
                    txt += f"\n{r.n_het} het"
            cell = ws.cell(row=row, column=col, value=txt)
            cell.fill = PatternFill("solid", fgColor=fills.get(r.verdict, "EEEEEE"))
            cell.alignment = center
            cell.font = Font(size=8)
            if best.get(r.called) is r:
                cell.border = gborder
        for col in range(1, 14):
            ws.column_dimensions[get_column_letter(col)].width = 11
        for row in range(1, 10):
            ws.row_dimensions[row].height = 38

    ws2 = wb.create_sheet("Summary")
    ws2.append(["plate", "well", "sample", "construct", "verdict", "reason",
                "mapped_reads", "breadth", "mean_depth", "mutations",
                "heterogeneous", "also_present"])
    for r in sorted(results, key=lambda x: (x.plate, x.well)):
        ws2.append([r.plate, r.well, r.sample, r.called or "", r.verdict, r.reason,
                    r.mapped_reads, round(r.breadth, 3), round(r.mean_depth, 1),
                    r.n_mut, r.n_het, r.runner_up or ""])
    for i, w in enumerate([10, 6, 32, 14, 22, 42, 12, 9, 10, 10, 13, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws3 = wb.create_sheet("Reports")
    mono = Font(name="Menlo", size=10)
    rn = 1
    for r in sorted(results, key=lambda x: (x.plate, x.well)):
        for line in text_block(r).split("\n"):
            ws3.cell(row=rn, column=1, value=line).font = mono
            rn += 1
        rn += 1
    ws3.column_dimensions["A"].width = 100
    wb.save(str(path))


# ─────────────────────────────────────────────────────────────────────────────
# 11.  Engine — reusable run() + parallel well loop
#
# The per-well analysis is embarrassingly parallel: evaluate_well touches only its
# own WellResult, and refs/index/distinct_maps are read-only. We fan the wells out
# across CPU cores with a spawn Pool. Two things the workers MUST receive: the big
# read-only `index` (shipped once per worker via the Pool initializer, never per
# task) and ERROR_RATE (a module global set at runtime — spawn workers re-import the
# module and would otherwise reset it to the default, silently changing --error-rate
# results). run() is importable so the web server can drive it directly with a
# progress callback; main() is a thin CLI wrapper that keeps identical stdout.
# ─────────────────────────────────────────────────────────────────────────────

class PickmeInputError(ValueError):
    """A folder that can't be analysed (no FASTQs / no maps / missing). Raised by
    run() so a caller (CLI or server) can report it instead of the process dying."""


# Worker-side cache of the shared read-only inputs (populated once per process by
# the Pool initializer, then read by every task in that worker).
_WORKER: dict = {}


def _init_worker(refs, index, distinct_maps, cap, error_rate):
    global ERROR_RATE
    ERROR_RATE = error_rate          # re-set the runtime global under spawn (the trap)
    _WORKER["refs"] = refs
    _WORKER["index"] = index
    _WORKER["distinct_maps"] = distinct_maps
    _WORKER["cap"] = cap


def _eval_one(task):
    """Analyse one well. task = (index, r1_path_str, r2_path_str_or_None, banned).
    Returns (index, WellResult) — the index lets the parent restore input order."""
    i, r1s, r2s, banned = task
    r1 = Path(r1s)
    r2 = Path(r2s) if r2s else None
    sample, plate, well = parse_sample(r1)
    res = WellResult(sample=sample, plate=plate, well=well)
    if r2 is None:
        res.verdict, res.reason = "NO_DATA", "no R2 mate"
        return i, res
    evaluate_well(res, r1, r2, _WORKER["refs"], _WORKER["index"],
                  _WORKER["distinct_maps"], _WORKER["cap"], banned=frozenset(banned))
    return i, res


def _run_wells(r1_sorted, refs, index, distinct_maps, cap, error_rate, jobs, progress,
               banned=()):
    """Analyse every well, in input order. jobs==1 runs serially in-process (the
    exact pre-parallel code path, for byte-identical verification); jobs>1 fans out
    across a spawn Pool and restores order via each task's index. `banned` = map names
    no well in THIS pass may be called as (see the second identity pass in run())."""
    n = len(r1_sorted)
    tasks = []
    for i, r1 in enumerate(r1_sorted):
        r2 = find_r2(r1)
        tasks.append((i, str(r1), (str(r2) if r2 else None), tuple(banned)))
    slots = [None] * n
    done = 0
    if jobs == 1:
        _init_worker(refs, index, distinct_maps, cap, error_rate)
        for t in tasks:
            i, res = _eval_one(t)
            slots[i] = res
            done += 1
            progress("analyzing", detail=res.sample, i=done, n=n)
    else:
        ctx = mp.get_context("spawn")
        with ctx.Pool(processes=jobs, initializer=_init_worker,
                      initargs=(refs, index, distinct_maps, cap, error_rate)) as pool:
            for i, res in pool.imap_unordered(_eval_one, tasks):
                slots[i] = res
                done += 1
                progress("analyzing", detail=res.sample, i=done, n=n)
    return slots


SF_DATALESS = 0x40000000   # macOS st_flags bit: a not-yet-materialised FileProvider file


def _is_dataless(p):
    """True if p is a not-yet-downloaded Google Drive / iCloud placeholder."""
    try:
        return bool(getattr(p.stat(), "st_flags", 0) & SF_DATALESS)
    except OSError:
        return True            # can't even stat it -> treat as needing a fetch


def ensure_local(paths, progress):
    """The "syncing" stage — download from Drive BEFORE analysing, and wait for it.

    On macOS a Google Drive / iCloud file that hasn't been downloaded yet is a
    'dataless' placeholder (SF_DATALESS in st_flags); its bytes are fetched from the
    cloud on first read. So a run must not just start reading — it must first pull down
    every input that isn't local yet, WAIT until each is fully here, and say so, or it
    would analyse half-present files and report false UNKNOWNs. (st_blocks is unreliable
    on the Drive mount — it caps ~4 MB — so we key off the dataless flag, not file size.)

    We find the placeholders, then download them CONCURRENTLY (this is I/O-bound waiting
    on Drive, so threads help a lot), reading each through to EOF to force a full
    materialisation, and report progress as "syncing" (N of M) so the user sees
    "Downloading from Drive …" instead of a stall. Files already on disk are skipped
    instantly. Tip: pinning the folder "Available offline" in Drive avoids the wait."""
    todo = [p for p in paths if _is_dataless(p)]
    m = len(todo)
    if not m:
        return                              # everything already local — nothing to sync

    progress("syncing", detail=f"downloading {m} file(s) from Drive", i=0, n=m)

    def fetch(p):
        try:
            with open(p, "rb") as f:        # reading to EOF pulls the whole file down
                while f.read(1 << 20):
                    pass
        except OSError:
            pass

    from concurrent.futures import ThreadPoolExecutor, as_completed
    done = 0
    with ThreadPoolExecutor(max_workers=min(8, m)) as ex:
        futs = {ex.submit(fetch, p): p for p in todo}
        for fut in as_completed(futs):
            done += 1
            progress("syncing", detail=futs[fut].name, i=done, n=m)

    # belt-and-suspenders: if anything is still dataless, wait briefly for Drive to settle
    for _ in range(20):
        if not any(_is_dataless(p) for p in todo):
            break
        for p in todo:
            if _is_dataless(p):
                fetch(p)

    # If files are STILL placeholders, this process could not pull them down (a headless
    # launchd/daemon context cannot trigger Google Drive's FileProvider — it returns
    # EDEADLK). Refuse to continue rather than silently analyse missing inputs and report
    # a plate of false UNKNOWNs. Be explicit and actionable about why and how to fix it.
    still = [p for p in todo if _is_dataless(p)]
    if still:
        shown = ", ".join(p.name for p in still[:4])
        more = f" (+{len(still) - 4} more)" if len(still) > 4 else ""
        raise PickmeInputError(
            f"could not download {len(still)} input file(s) from Google Drive: {shown}{more}. "
            f"These are un-downloaded 'online-only' placeholders and this server cannot fetch "
            f"them itself. Fix: in Google Drive, right-click the run folder → \"Available "
            f"offline\" (or open the files once in Finder) so Drive downloads them, then run "
            f"again. Analysis was NOT started — better no result than a wrong one.")


def run(folder, *, jobs=None, cap=20000, error_rate=0.01,
        fastq_dir=None, refs_dir=None, out=None, progress=None):
    """Analyse a folder end-to-end and write pickme_results/plate_map.{html,xlsx,txt,json}.
    Returns (results, out_dir). `progress(stage, *, detail="", i=None, n=None)` is an
    optional callback invoked at each phase boundary (stages: discover, syncing,
    reading, indexing, analyzing, crosscheck, writing, done); it defaults to a no-op so
    pickme stays server-agnostic. Raises PickmeInputError on unusable input."""
    def prog(stage, *, detail="", i=None, n=None):
        if progress:
            progress(stage, detail=detail, i=i, n=n)

    global ERROR_RATE
    ERROR_RATE = error_rate
    _ensure_deps()

    root = Path(folder).expanduser().resolve()
    if not root.exists():
        raise PickmeInputError(f"folder not found: {root}")
    fq = Path(fastq_dir).expanduser().resolve() if fastq_dir else None
    rf = Path(refs_dir).expanduser().resolve() if refs_dir else None

    print(f"Scanning: {root}")
    prog("discover", detail=str(root))
    ref_files, r1_files = discover(root, fq, rf)
    print(f"  {len(ref_files)} reference file(s), {len(r1_files)} FASTQ pair(s)")
    if not ref_files:
        raise PickmeInputError("no .gb/.dna references found.")
    if not r1_files:
        raise PickmeInputError("no *_R1*.fastq(.gz) files found.")
    r1_sorted = sorted(r1_files)

    # syncing — materialise online-only Drive placeholders before analysis stalls on them
    sync_paths = list(r1_sorted)
    sync_paths += [r for r in (find_r2(x) for x in r1_sorted) if r]
    sync_paths += list(ref_files)
    ensure_local(sync_paths, prog)

    print("Loading references ...")
    prog("reading", detail=f"{len(r1_files)} samples", n=len(r1_files))
    refs = load_references(ref_files)
    print(f"  {len(refs)} reference(s)")
    njobs = jobs if jobs else (os.cpu_count() or 1)
    print("Indexing & grouping ...")
    prog("indexing", detail=f"{len(refs)} references")
    index = build_index(refs)
    ref_by_group = group_refs(refs)
    distinct_maps = build_distinctive(refs, ref_by_group, jobs=njobs)
    print(f"  {len(ref_by_group)} construct group(s)")

    out_dir = (Path(out).expanduser().resolve() if out else root / "pickme_results")
    out_dir.mkdir(parents=True, exist_ok=True)

    results = _run_wells(r1_sorted, refs, index, distinct_maps, cap, error_rate,
                         njobs, prog)

    # ── plate-wide pass: demote map errors so good clones aren't condemned ──
    print("Cross-checking the plate (separating map errors from real mutations) ...")
    prog("crosscheck", detail="separating map errors from real mutations")

    # Which constructs are plate-wide background? Only the PLATE can tell us — one well
    # cannot distinguish "my clone" from "a smear everyone has". Wells whose call is just
    # that smear get a SECOND identity pass with it suppressed, so the well's real clone
    # is finally scored instead of the well being thrown away. What the smear was, and at
    # what depth, is kept on the result as evidence for the reader to judge.
    bg = find_background_maps(results)
    if bg:
        redo = shallow_background_wells(results, bg)
        names = ", ".join(sorted(n.split("__")[0] for n in bg))
        print(f"  plate-wide background: {names}")
        if redo:
            prog("crosscheck", detail=f"re-identifying {len(redo)} well(s) without background")
            print(f"  re-identifying {len(redo)} well(s) with the background suppressed ...")
            sub = [r1_sorted[i] for i in redo]
            again = _run_wells(sub, refs, index, distinct_maps, cap, error_rate,
                               njobs, prog, banned=tuple(bg))
            recovered = 0
            for i, new in zip(redo, again):
                old = results[i]
                new.background = [(old.called, round(old.mean_depth, 1),
                                   round(bg[old.called][0], 1))]
                if new.called:
                    recovered += 1
                else:
                    # nothing else in the well either — say so honestly, naming the smear
                    c, d, md = new.background[0]
                    new.reason = (f"only {c.split('__')[0]} found, and only as background "
                                  f"({d:.0f}x here vs {md:.0f}x where it is real) — "
                                  f"no other construct's own sequence is present")
                results[i] = new
            print(f"  recovered {recovered} of {len(redo)} well(s) hidden under background")
    mark_systematic(results, refs)

    # ── corrected .gb per confidently-called well ──
    for res in results:
        if (res.called_idx >= 0 and not res.mix_flag
                and res.verdict not in ("RED_MIX", "RED_EMPTY", "RED_UNKNOWN", "NO_DATA")):
            try:
                res.corrected_gb = build_corrected_gb(res, refs[res.called_idx])
            except Exception as e:
                res.corrected_gb = ""
                print(f"  ! corrected .gb for {res.well} failed ({e})")

    for i, res in enumerate(results, 1):
        extra = (f"  {res.n_mut}mut/{res.n_het}het/{res.n_systematic}sys"
                 if res.called else "")
        if res.runner_up:
            extra += f"  +{res.runner_up}"
        print(f"  [{i}/{len(results)}] {res.sample:38} {res.verdict:11} "
              f"{res.called or '-':12} {res.breadth:.0%} {res.mean_depth:.0f}x{extra}")

    base = out_dir / "plate_map"
    prog("writing", detail="plate_map.html + xlsx/txt/json")
    write_html(results, base.with_suffix(".html"))
    write_xlsx(results, base.with_suffix(".xlsx"))
    write_text(results, base.with_suffix(".txt"))
    write_json(results, base.with_suffix(".json"))
    print(f"\nDone.\n  {base.with_suffix('.html')}   <- open this\n"
          f"  {base.with_suffix('.xlsx')}\n  {base.with_suffix('.txt')}"
          f"\n  {base.with_suffix('.json')}")
    counts = Counter(r.verdict for r in results)
    print("\nSummary:")
    for v in ("GREEN", "YELLOW", "RED_MIX", "RED_MUT", "RED_HET", "RED_BROKEN",
              "RED_EMPTY", "RED_UNKNOWN", "NO_DATA"):
        if counts.get(v):
            print(f"  {v:12} {counts[v]}")
    prog("done", detail=str(base.with_suffix(".html")))
    return results, out_dir


# ─────────────────────────────────────────────────────────────────────────────
# 12.  Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        prog="pickme.py",
        description="pickme — 'can I pick this clone?'. Drop this in a folder of "
                    "paired FASTQs + .gb reference maps and run; get a colored "
                    "plate map of which construct is in each well and whether "
                    "it's clean. Pure Python, runs on any Mac/Windows.")
    ap.add_argument("folder", nargs="?", default=".", help="Folder to scan.")
    ap.add_argument("--fastq", help="FASTQ folder (default: auto-detect).")
    ap.add_argument("--refs", help="Reference folder (default: auto-detect).")
    ap.add_argument("--out", help="Output folder (default: <folder>/pickme_results).")
    ap.add_argument("--cap", type=int, default=20000,
                    help="Max reads per R1/R2 file (speed cap; default 20000).")
    ap.add_argument("--error-rate", type=float, default=0.01,
                    help="Assumed per-base sequencing error for the mutation test "
                         "(default 0.01 = 1%%). Lower it for high-quality data.")
    ap.add_argument("-j", "--jobs", type=int, default=None,
                    help="Parallel worker processes (default: all CPU cores; "
                         "1 = serial, the pre-parallel code path).")
    ap.add_argument("--no-open", action="store_true",
                    help="Don't auto-open the HTML report when finished.")
    args = ap.parse_args()

    try:
        _results, out_dir = run(
            args.folder, jobs=args.jobs, cap=args.cap, error_rate=args.error_rate,
            fastq_dir=args.fastq, refs_dir=args.refs, out=args.out)
    except PickmeInputError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)

    # Open the report in the default browser so "run it -> see it" is one step.
    if not args.no_open:
        try:
            import webbrowser
            webbrowser.open((out_dir / "plate_map.html").as_uri())
        except Exception:
            pass  # headless / no browser — the file is still written.


if __name__ == "__main__":
    main()
