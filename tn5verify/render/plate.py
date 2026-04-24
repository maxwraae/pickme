from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

from ..types import WellResult
from .well_report import _format_block


def _blend(hex1: str, hex2: str, t: float) -> str:
    """Blend two hex colours; t in [0, 1]. Returns 6-char hex (no #)."""
    r1, g1, b1 = int(hex1[1:3], 16), int(hex1[3:5], 16), int(hex1[5:7], 16)
    r2, g2, b2 = int(hex2[1:3], 16), int(hex2[3:5], 16), int(hex2[5:7], 16)
    r = int(r1 + (r2 - r1) * t)
    g = int(g1 + (g2 - g1) * t)
    b = int(b1 + (b2 - b1) * t)
    return f"{r:02X}{g:02X}{b:02X}"


def _mean_cov(r: WellResult) -> float:
    if not r.backbone_windows:
        return 0.0
    return sum(w.mean_coverage for w in r.backbone_windows) / len(r.backbone_windows)


def _fill_color(r: WellResult, mean_cov: float) -> str:
    """Return 6-char hex fill colour for a well."""
    verdict = r.verdict
    intensity = min(1.0, mean_cov / 200.0)

    if verdict == "GREEN":
        return _blend("#C8E6C9", "#1B5E20", intensity)
    elif verdict == "YELLOW":
        return _blend("#FFF9C4", "#F57F17", intensity)
    elif verdict.startswith("RED"):
        return _blend("#FFCDD2", "#B71C1C", intensity)
    else:  # NO_DATA, RED_UNKNOWN, anything else
        return "EEEEEE"


def _parse_well_id(well_id: str):
    """Parse e.g. 'B4' → (row_idx 2, col_idx 4) 1-based for Excel.
    Returns (None, None) if unparseable."""
    if not well_id or len(well_id) < 2:
        return None, None
    row_letter = well_id[0].upper()
    if row_letter not in "ABCDEFGH":
        return None, None
    try:
        col_num = int(well_id[1:])
    except ValueError:
        return None, None
    if col_num < 1 or col_num > 12:
        return None, None
    row_idx = ord(row_letter) - ord("A") + 1  # A=1 … H=8
    return row_idx, col_num


_THICK_GOLD = Side(style="thick", color="FFD700")
_BEST_BORDER = Border(
    left=_THICK_GOLD,
    right=_THICK_GOLD,
    top=_THICK_GOLD,
    bottom=_THICK_GOLD,
)


def write_xlsx(results: list[WellResult], path: Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ── Plate Map sheet ────────────────────────────────────────────────────────
    ws_plate = wb.active
    ws_plate.title = "Plate Map"

    # Header row: columns 1–12
    for col in range(1, 13):
        cell = ws_plate.cell(row=1, column=col + 1, value=col)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row labels A–H
    row_letters = list("ABCDEFGH")
    for i, letter in enumerate(row_letters):
        cell = ws_plate.cell(row=i + 2, column=1, value=letter)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Compute best picks: called_member → well with highest mean_cov & GREEN
    best_picks: dict[str, WellResult] = {}
    for r in results:
        if r.called_member and r.verdict == "GREEN":
            cov = _mean_cov(r)
            existing = best_picks.get(r.called_member)
            if existing is None or cov > _mean_cov(existing):
                best_picks[r.called_member] = r

    # Fill in wells
    for r in results:
        row_idx, col_idx = _parse_well_id(r.well_id)
        if row_idx is None:
            continue

        cov = _mean_cov(r)
        hex_color = _fill_color(r, cov)
        fill = PatternFill(patternType="solid", fgColor=hex_color)

        called = r.called_member or "?"
        if r.variant_regions:
            clean_count = sum(
                1 for rr in r.variant_regions
                for pos in rr.positions
                if pos.get("status") in ("CLEAN", "THIN")
            )
            total_pos = sum(len(rr.positions) for rr in r.variant_regions)
            content = f"{called}\n{cov:.0f}×  {clean_count}/{total_pos}"
        else:
            content = f"{called}\n{cov:.0f}×"

        # Excel row: header is row 1, so well rows start at 2; row A → row 2
        excel_row = row_idx + 1
        excel_col = col_idx + 1

        cell = ws_plate.cell(row=excel_row, column=excel_col, value=content)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

        # Best pick border
        if r.called_member and best_picks.get(r.called_member) is r:
            cell.border = _BEST_BORDER

    # Column widths and row heights
    for col in range(1, 14):
        ws_plate.column_dimensions[get_column_letter(col)].width = 12
    for row in range(1, 10):
        ws_plate.row_dimensions[row].height = 30

    # ── Picks sheet ───────────────────────────────────────────────────────────
    ws_picks = wb.create_sheet("Picks")
    ws_picks.append(["Construct", "Well"])
    for construct, r in sorted(best_picks.items()):
        ws_picks.append([construct, r.well_id])

    # ── Well Reports sheet ───────────────────────────────────────────────────
    # Full per-well analysis: why we called each construct, raw pileup counts,
    # backbone profile, verdict reason, suggested action. Monospaced so the
    # per-position A/T/C/G tables line up.
    ws_reports = wb.create_sheet("Well Reports")
    monospace = Font(name="Menlo", size=10)

    def _row_order(r: WellResult) -> tuple[int, int]:
        row_idx, col_idx = _parse_well_id(r.well_id)
        return (row_idx or 99, col_idx or 99)

    row_num = 1
    for result in sorted(results, key=_row_order):
        block = _format_block(result)
        for line in block.split("\n"):
            cell = ws_reports.cell(row=row_num, column=1, value=line)
            cell.font = monospace
            cell.alignment = Alignment(vertical="top")
            row_num += 1
        # Divider between wells
        divider_cell = ws_reports.cell(row=row_num, column=1, value="─" * 60)
        divider_cell.font = monospace
        row_num += 2  # blank spacer row too

    ws_reports.column_dimensions["A"].width = 110

    wb.save(str(path))
